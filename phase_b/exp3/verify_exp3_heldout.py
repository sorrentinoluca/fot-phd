#!/usr/bin/env python3
"""Fail-closed structural/provenance verifier for the Experiment 3 held-out."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import subprocess
import sys
from collections import Counter
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from jsonschema import Draft202012Validator, FormatChecker
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
EXPECTED_CONDITIONS = ("Normal", "F1", "F8", "F10", "F13")
EXPECTED_HEADER = (
    ["Time (h)"]
    + [f"XMEAS-{index}" for index in range(1, 42)]
    + [f"XMV-{index}" for index in range(1, 13)]
)
EXPECTED_RUNTIME = {
    "matlab_version_full": "25.2.0.3312555 (R2025b) Update 6",
    "matlab_release": "2025b",
    "matlab_build": "3312555",
    "matlab_product_date": "28-Jul-2025",
    "matlab_runtime_update_date": "June 30, 2026",
    "simulink_version": "25.2",
    "simulink_release": "(R2025b)",
    "simulink_product_date": "28-Jul-2025",
    "architecture": "MACA64",
    "matlabroot": "/Applications/MATLAB_R2025b.app",
}
EXPECTED_SIMULATOR = {
    "simulator_commit": "a0413e16c940f0fc8b554d6a86248020d7fb7527",
    "model_name": "MultiLoop_mode1",
    "simulation_mode": "normal",
    "solver": "ode45",
    "sfunction_identity": "temexd_mod",
    "sfunction_hash": "0da41d939e5ab7ba122d7b70c124368ee0882fce40e775dba5d180e7a7e24e5e",
    "sfunction_mex_hash": "68f632388cb698dd7b8c595000bc03c2e1d19200546b9d4357df90e3fc93af0d",
    "model_hash": "d2f6659f65935021d4b1813e7189be02e7ae9f5639b794e8edc4f2f3c5cddba8",
    "initial_state_hash": "40eaebc92badb04ad026e358cfd28ec9c778fcf2d24a1b8f5d85565854da2747",
}
ATTEMPT_REQUIRED_FIELDS = {
    "physical_case_id",
    "fault_status",
    "attempt",
    "seed",
    "rng_algorithm",
    "started_at",
    "completed_at",
    *EXPECTED_RUNTIME,
    *EXPECTED_SIMULATOR,
    "case_plan_hash",
    "generation_script_hash",
    "output_path",
    "output_size_bytes",
    "output_sha256",
    "rows",
    "cols",
    "time_start",
    "time_end",
    "sampling_interval",
    "finite_check",
    "structural_valid",
    "technical_failure_reason",
}
MANIFEST_FIELDS = (
    "physical_case_id",
    "fault/status",
    "attempt",
    "seed",
    "filename",
    "size_bytes",
    "SHA256",
    "rows",
    "cols",
    "time_start",
    "time_end",
    "sampling",
    "finite_check",
    "structural_valid",
)
TIME_TOLERANCE = 1e-10
HASH_PATTERN = re.compile(r"^[0-9a-f]{64}$")
EXPECTED_FREEZE_PATHS = {
    ".gitignore",
    "requirements.txt",
    "phase_b/exp3/EXP3_FRESH_RUN_PROTOCOL.md",
    "phase_b/exp3/RNG_RUNTIME_VALIDATION.md",
    "phase_b/exp3/exp3_attempt_log.schema.json",
    "phase_b/exp3/exp3_attempt_log.template.json",
    "phase_b/exp3/exp3_case_plan.json",
    "phase_b/exp3/exp3_manifest_template.csv",
    "phase_b/exp3/generate_exp3_heldout.m",
    "phase_b/exp3/validate_exp3_rng_runtime.m",
    "phase_b/exp3/verify_exp3_heldout.py",
    "phase_b/tests/test_exp3_pre_freeze.py",
}
EXP3_FREEZE_TAG = "exp3-heldout-frozen"
EXP3_FREEZE_COMMIT = "b02e93f92bf6fa85a4fd0a2e010bac365a3a7c89"
HOTFIX_001_COMMIT = "cdba0202435d1c97ea79cfff586e59534ce9baad"
HOTFIX_001_PATH = SCRIPT_DIR / "EXP3_POST_FREEZE_HOTFIX_001.json"
HOTFIX_002_COMMIT = "28130023a34eda778c04a001a9f631404bd6b9a6"
HOTFIX_002_TAG = "exp3-post-freeze-hotfix-002"
HOTFIX_002_PATH = SCRIPT_DIR / "EXP3_POST_FREEZE_HOTFIX_002.json"
HOTFIX_003_PATH = SCRIPT_DIR / "EXP3_POST_FREEZE_HOTFIX_003.json"
HOTFIX_002_ARTIFACT_PATHS = {
    "phase_b/exp3/EXP3_FRESH_RUN_PROTOCOL.md",
    "phase_b/exp3/EXP3_POST_FREEZE_HOTFIX_002.md",
    "phase_b/exp3/RNG_RUNTIME_VALIDATION.md",
    "phase_b/exp3/exp3_attempt_log.schema.json",
    "phase_b/exp3/exp3_case_plan.json",
    "phase_b/exp3/generate_exp3_heldout.m",
    "phase_b/exp3/test_exp3_runtime_provenance.m",
    "phase_b/exp3/validate_exp3_rng_runtime.m",
    "phase_b/exp3/verify_exp3_heldout.py",
    "phase_b/tests/test_exp3_pre_freeze.py",
}
HOTFIX_003_ARTIFACT_PATHS = {
    "phase_b/exp3/EXP3_FRESH_RUN_PROTOCOL.md",
    "phase_b/exp3/EXP3_POST_FREEZE_HOTFIX_003.md",
    "phase_b/exp3/generate_exp3_heldout.m",
    "phase_b/exp3/restore_exp3_plot_stopfcn.m",
    "phase_b/exp3/suppress_exp3_plot_stopfcn.m",
    "phase_b/exp3/test_exp3_attempt_policy.m",
    "phase_b/exp3/test_exp3_stopfcn_management.m",
    "phase_b/exp3/verify_exp3_heldout.py",
    "phase_b/tests/test_exp3_pre_freeze.py",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sha256_bytes(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def git_show_bytes(revision: str, relative: str) -> bytes:
    return subprocess.run(
        ["git", "show", f"{revision}:{relative}"],
        cwd=ROOT,
        check=True,
        capture_output=True,
    ).stdout


def close(left: float, right: float) -> bool:
    return math.isclose(left, right, rel_tol=0.0, abs_tol=TIME_TOLERANCE)


def parse_bool(value: Any, field: str) -> bool:
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized not in {"true", "false"}:
        raise ValueError(f"{field} must be true or false, found {value!r}")
    return normalized == "true"


def populated(value: Any) -> bool:
    return value is not None and bool(str(value).strip())


def canonical_cases() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    ordinal = 0
    for condition in EXPECTED_CONDITIONS:
        prefix = "N" if condition == "Normal" else condition
        for run_index in range(1, 7):
            ordinal += 1
            primary_seed = 310000 + ordinal
            rows.append(
                {
                    "physical_case_id": f"EXP3-{prefix}-{run_index:03d}",
                    "condition": condition,
                    "run_index": run_index,
                    "rng_algorithm": "twister",
                    "primary_seed": primary_seed,
                    "replacement_seed": primary_seed + 1_000_000,
                    "max_total_attempts": 2,
                }
            )
    return rows


def load_case_plan(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    errors = validate_case_plan(payload)
    if errors:
        raise ValueError("invalid case plan: " + "; ".join(errors))
    return payload


def validate_freeze_manifest(path: Path) -> list[str]:
    errors: list[str] = []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"freeze manifest cannot be read: {exc}"]
    expected_metadata = {
        "schema_version": "1.0",
        "experiment": "Experiment 3 — Fresh Prospective Physical-Run Extension",
        "status": "FROZEN_BEFORE_GENERATION",
        "hash_algorithm": "SHA-256",
        "freeze_tag": "exp3-heldout-frozen",
        "source_head_before_freeze": "430590001922b28d618b739b12e3471e7ebd0afa",
        "created_before_generation": True,
        "exp3_workbooks_at_freeze": 0,
    }
    for field, expected in expected_metadata.items():
        if payload.get(field) != expected:
            errors.append(f"freeze manifest {field} mismatch")
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, list):
        return errors + ["freeze manifest artifacts must be an array"]
    paths = [row.get("path") for row in artifacts if isinstance(row, dict)]
    if len(paths) != len(set(paths)):
        errors.append("freeze manifest contains duplicate paths")
    if set(paths) != EXPECTED_FREEZE_PATHS:
        errors.append("freeze manifest artifact path set mismatch")
    try:
        freeze_target = subprocess.run(
            ["git", "rev-parse", f"{EXP3_FREEZE_TAG}^{{}}"],
            cwd=ROOT,
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()
        if freeze_target != EXP3_FREEZE_COMMIT:
            errors.append("original freeze tag target mismatch")
        tagged_manifest = git_show_bytes(
            EXP3_FREEZE_TAG, "phase_b/exp3/EXP3_FREEZE_MANIFEST.json"
        )
        if path.read_bytes() != tagged_manifest:
            errors.append("original freeze manifest differs from tagged bytes")
    except Exception as exc:
        errors.append(f"original freeze tag cannot be verified: {exc}")

    for row in artifacts:
        if not isinstance(row, dict) or set(row) != {"path", "sha256"}:
            errors.append("freeze manifest artifact entry schema mismatch")
            continue
        relative = Path(row["path"])
        if relative.is_absolute() or ".." in relative.parts:
            errors.append(f"invalid freeze-manifest path: {row['path']!r}")
            continue
        expected_hash = row["sha256"]
        if not isinstance(expected_hash, str) or not HASH_PATTERN.fullmatch(
            expected_hash
        ):
            errors.append(f"invalid frozen SHA-256: {row['path']}")
            continue
        try:
            tagged_bytes = git_show_bytes(EXP3_FREEZE_TAG, row["path"])
        except Exception as exc:
            errors.append(f"frozen tagged artifact missing: {row['path']}: {exc}")
            continue
        if sha256_bytes(tagged_bytes) != expected_hash:
            errors.append(f"frozen tagged artifact hash mismatch: {row['path']}")

    errors.extend(validate_hotfix_002(path))
    errors.extend(validate_hotfix_003(path))
    return errors


def validate_hotfix_002(freeze_manifest_path: Path) -> list[str]:
    errors: list[str] = []
    try:
        hotfix = json.loads(HOTFIX_002_PATH.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"hotfix 002 manifest cannot be read: {exc}"]
    expected_hotfix_metadata = {
        "schema_version": "1.0",
        "hotfix_id": "EXP3_POST_FREEZE_HOTFIX_002",
        "status": "AUTHORIZED_BEFORE_FIRST_SIMULATION",
        "original_freeze_commit": EXP3_FREEZE_COMMIT,
        "original_freeze_tag": EXP3_FREEZE_TAG,
        "hotfix_001_commit": HOTFIX_001_COMMIT,
        "scientific_protocol_changed": False,
        "experiment_1_frozen_artifacts_changed": False,
    }
    for field, expected in expected_hotfix_metadata.items():
        if hotfix.get(field) != expected:
            errors.append(f"hotfix 002 {field} mismatch")
    if hotfix.get("original_freeze_manifest_sha256") != sha256_file(
        freeze_manifest_path
    ):
        errors.append("hotfix 002 original freeze manifest hash mismatch")
    if hotfix.get("hotfix_001_manifest_sha256") != sha256_file(HOTFIX_001_PATH):
        errors.append("hotfix 002 hotfix 001 manifest hash mismatch")
    try:
        hotfix001 = json.loads(HOTFIX_001_PATH.read_text(encoding="utf-8"))
        hotfix001_generator = sha256_bytes(
            git_show_bytes(HOTFIX_001_COMMIT, "phase_b/exp3/generate_exp3_heldout.m")
        )
        if (
            hotfix.get("hotfix_001_generator_sha256") != hotfix001_generator
            or hotfix001.get("hotfixed_generator_sha256") != hotfix001_generator
        ):
            errors.append("hotfix 002 hotfix 001 generator hash mismatch")
    except Exception as exc:
        errors.append(f"hotfix 001 chain cannot be verified: {exc}")

    expected_boundary = {
        "physical_case_id": "EXP3-N-001",
        "attempt": 0,
        "seed": 310001,
        "sim_called": False,
        "run_rng_consumed": False,
        "output_directories_created": True,
        "directories_empty": True,
        "workbooks_created": 0,
        "attempt_log_created": False,
        "final_manifest_created": False,
        "scientific_outcome_observed": False,
    }
    if hotfix.get("failed_invocation") != expected_boundary:
        errors.append("hotfix 002 failed-invocation boundary mismatch")
    expected_semantics = {
        "matlab_version_full": "version",
        "matlab_release": "version('-release')",
        "matlab_build": "parsed from version",
        "matlab_product_date": "ver('MATLAB').Date",
        "matlab_runtime_update_date": "version('-date')",
    }
    if hotfix.get("runtime_semantics") != expected_semantics:
        errors.append("hotfix 002 runtime field semantics mismatch")

    try:
        target = subprocess.run(
            ["git", "rev-parse", f"{HOTFIX_002_TAG}^{{}}"],
            cwd=ROOT,
            check=True,
            capture_output=True,
            text=True,
        ).stdout.strip()
        if target != HOTFIX_002_COMMIT:
            errors.append("hotfix 002 tag target mismatch")
        if HOTFIX_002_PATH.read_bytes() != git_show_bytes(
            HOTFIX_002_COMMIT,
            "phase_b/exp3/EXP3_POST_FREEZE_HOTFIX_002.json",
        ):
            errors.append("hotfix 002 manifest differs from committed bytes")
    except Exception as exc:
        errors.append(f"hotfix 002 commit/tag cannot be verified: {exc}")

    changed = hotfix.get("changed_artifacts")
    errors.extend(
        validate_hotfix_artifacts(
            changed,
            HOTFIX_002_ARTIFACT_PATHS,
            HOTFIX_001_COMMIT,
            HOTFIX_002_COMMIT,
            "hotfix 002",
        )
    )
    artifact_after = {
        row.get("path"): row.get("after_sha256")
        for row in changed or []
        if isinstance(row, dict)
    }
    if hotfix.get("hotfix_002_generator_sha256") != artifact_after.get(
        "phase_b/exp3/generate_exp3_heldout.m"
    ):
        errors.append("hotfix 002 generator hash aliases disagree")
    if hotfix.get("hotfix_002_case_plan_sha256") != artifact_after.get(
        "phase_b/exp3/exp3_case_plan.json"
    ):
        errors.append("hotfix 002 case-plan hash aliases disagree")
    return errors


def validate_hotfix_003(freeze_manifest_path: Path) -> list[str]:
    errors: list[str] = []
    try:
        hotfix = json.loads(HOTFIX_003_PATH.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"hotfix 003 manifest cannot be read: {exc}"]
    expected_metadata = {
        "schema_version": "1.0",
        "hotfix_id": "EXP3_POST_FREEZE_HOTFIX_003",
        "status": "AUTHORIZED_BEFORE_REPLACEMENT_ATTEMPT",
        "original_freeze_commit": EXP3_FREEZE_COMMIT,
        "original_freeze_tag": EXP3_FREEZE_TAG,
        "hotfix_001_commit": HOTFIX_001_COMMIT,
        "hotfix_002_commit": HOTFIX_002_COMMIT,
        "hotfix_002_tag": HOTFIX_002_TAG,
        "scientific_protocol_changed": False,
        "experiment_1_frozen_artifacts_changed": False,
    }
    for field, expected in expected_metadata.items():
        if hotfix.get(field) != expected:
            errors.append(f"hotfix 003 {field} mismatch")
    if hotfix.get("original_freeze_manifest_sha256") != sha256_file(
        freeze_manifest_path
    ):
        errors.append("hotfix 003 original freeze manifest hash mismatch")
    if hotfix.get("hotfix_001_manifest_sha256") != sha256_file(HOTFIX_001_PATH):
        errors.append("hotfix 003 hotfix 001 manifest hash mismatch")
    if hotfix.get("hotfix_002_manifest_sha256") != sha256_file(HOTFIX_002_PATH):
        errors.append("hotfix 003 hotfix 002 manifest hash mismatch")
    expected_boundary = {
        "physical_case_id": "EXP3-N-001",
        "attempt": 0,
        "seed": 310001,
        "sim_called": True,
        "sim_returned_successfully": False,
        "workbooks_created": 0,
        "output_size_bytes": 0,
        "output_sha256": "",
        "signal_inspection": False,
        "attempt_log_created": True,
        "attempt_0_recorded_technical_failure": True,
        "scientific_accepted_output": False,
    }
    if hotfix.get("failed_invocation") != expected_boundary:
        errors.append("hotfix 003 failed-invocation boundary mismatch")
    expected_restart = {
        "physical_case_id": "EXP3-N-001",
        "attempt": 1,
        "seed": 1310001,
        "rng_algorithm": "twister",
    }
    if hotfix.get("authorized_restart") != expected_restart:
        errors.append("hotfix 003 authorized-restart mismatch")
    expected_callback = {
        "expected_original_stopfcn": "TEplot",
        "temporary_stopfcn": "",
        "restore_required": True,
        "model_save_allowed": False,
        "model_sha256": EXPECTED_SIMULATOR["model_hash"],
    }
    if hotfix.get("callback_management") != expected_callback:
        errors.append("hotfix 003 callback-management mismatch")
    if hotfix.get("attempt_log_sha256_at_discovery") != (
        "0b2f2e6bf3c82e85da72591919fade41033c63431202f2f97dae6bd1d59a9729"
    ):
        errors.append("hotfix 003 attempt-log discovery hash mismatch")
    changed = hotfix.get("changed_artifacts")
    errors.extend(
        validate_hotfix_artifacts(
            changed,
            HOTFIX_003_ARTIFACT_PATHS,
            HOTFIX_002_COMMIT,
            None,
            "hotfix 003",
        )
    )
    artifact_after = {
        row.get("path"): row.get("after_sha256")
        for row in changed or []
        if isinstance(row, dict)
    }
    if hotfix.get("hotfix_003_generator_sha256") != artifact_after.get(
        "phase_b/exp3/generate_exp3_heldout.m"
    ):
        errors.append("hotfix 003 generator hash aliases disagree")
    if hotfix.get("hotfix_003_case_plan_sha256") != sha256_file(
        SCRIPT_DIR / "exp3_case_plan.json"
    ):
        errors.append("hotfix 003 case-plan hash mismatch")
    return errors


def validate_hotfix_artifacts(
    changed: Any,
    expected_paths: set[str],
    before_revision: str,
    after_revision: str | None,
    label: str,
) -> list[str]:
    errors: list[str] = []
    if not isinstance(changed, list):
        return [f"{label} changed_artifacts must be an array"]
    changed_paths = [row.get("path") for row in changed if isinstance(row, dict)]
    if len(changed_paths) != len(set(changed_paths)):
        errors.append(f"{label} contains duplicate artifact paths")
    if set(changed_paths) != expected_paths:
        errors.append(f"{label} artifact path set mismatch")
    for row in changed:
        if not isinstance(row, dict) or set(row) != {
            "path",
            "before_sha256",
            "after_sha256",
        }:
            errors.append(f"{label} artifact entry schema mismatch")
            continue
        relative = Path(row["path"])
        if relative.is_absolute() or ".." in relative.parts:
            errors.append(f"invalid {label} path: {row['path']!r}")
            continue
        after_hash = row["after_sha256"]
        if not isinstance(after_hash, str) or not HASH_PATTERN.fullmatch(after_hash):
            errors.append(f"invalid {label} SHA-256: {row['path']}")
            continue
        try:
            if after_revision is None:
                artifact = ROOT / relative
                if not artifact.is_file():
                    errors.append(f"{label} artifact missing: {row['path']}")
                    continue
                observed_after = sha256_file(artifact)
            else:
                observed_after = sha256_bytes(
                    git_show_bytes(after_revision, row["path"])
                )
            if observed_after != after_hash:
                errors.append(f"{label} artifact hash mismatch: {row['path']}")
        except Exception as exc:
            errors.append(f"{label} artifact cannot be verified: {row['path']}: {exc}")
        before_hash = row["before_sha256"]
        try:
            previous_bytes = git_show_bytes(before_revision, row["path"])
        except subprocess.CalledProcessError:
            if before_hash is not None:
                errors.append(f"{label} new artifact has prior hash: {row['path']}")
        else:
            if (
                not isinstance(before_hash, str)
                or not HASH_PATTERN.fullmatch(before_hash)
                or sha256_bytes(previous_bytes) != before_hash
            ):
                errors.append(f"{label} prior artifact hash mismatch: {row['path']}")
    return errors


def validate_case_plan(plan: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if plan.get("status") not in {
        "PRE_FREEZE_DRAFT",
        "FROZEN_BEFORE_GENERATION",
    }:
        errors.append("case-plan status is neither pre-freeze draft nor frozen")
    expected = canonical_cases()
    cases = plan.get("cases")
    if not isinstance(cases, list):
        return ["cases must be an array"]
    if cases != expected:
        errors.append("cases do not exactly match the deterministic 30-case plan")

    ids = [row.get("physical_case_id") for row in cases]
    primary = [row.get("primary_seed") for row in cases]
    replacement = [row.get("replacement_seed") for row in cases]
    if len(cases) != 30:
        errors.append(f"expected 30 cases, found {len(cases)}")
    if len(set(ids)) != len(ids):
        errors.append("physical_case_id values are not unique")
    if len(set(primary)) != len(primary):
        errors.append("primary seeds are not unique")
    if len(set(replacement)) != len(replacement):
        errors.append("replacement seeds are not unique")
    if set(primary) & set(replacement):
        errors.append("primary/replacement seed collision")
    if any(r != p + 1_000_000 for p, r in zip(primary, replacement)):
        errors.append("replacement seed formula mismatch")
    counts = Counter(row.get("condition") for row in cases)
    if counts != Counter({condition: 6 for condition in EXPECTED_CONDITIONS}):
        errors.append(f"condition counts mismatch: {dict(counts)}")

    rng = plan.get("rng", {})
    required_rng = {
        "algorithm": "twister",
        "master_allocation_base": 310000,
        "replacement_seed_formula": "primary_seed + 1000000",
        "allowed_attempts": [0, 1],
        "max_total_attempts": 2,
        "bootstrap_seed": 310031,
    }
    for field, expected_value in required_rng.items():
        if rng.get(field) != expected_value:
            errors.append(f"rng.{field} mismatch")

    sample = plan.get("sample", {})
    if sample.get("total_physical_runs") != 30:
        errors.append("sample.total_physical_runs must equal 30")
    if sample.get("total_fault_physical_runs") != 24:
        errors.append("sample.total_fault_physical_runs must equal 24")

    simulator = plan.get("simulator", {})
    simulator_expected = {
        "simulator_commit": EXPECTED_SIMULATOR["simulator_commit"],
        "model": EXPECTED_SIMULATOR["model_name"],
        "tep_mode": "Mode 1",
        "simulation_mode": EXPECTED_SIMULATOR["simulation_mode"],
        "solver": EXPECTED_SIMULATOR["solver"],
        "start_time_h": 0.0,
        "stop_time_h": 50.0,
        "sampling_interval_h": 1 / 60,
        "fault_injection_h": 10.0,
        "initial_state_file": "Mode1xInitial.mat",
        "sfunction_identity": EXPECTED_SIMULATOR["sfunction_identity"],
        "custom_setpoints": False,
        "expected_rows": 3001,
        "expected_columns": 54,
    }
    for field, expected_value in simulator_expected.items():
        observed = simulator.get(field)
        if isinstance(expected_value, float):
            if not isinstance(observed, (int, float)) or not close(
                float(observed), expected_value
            ):
                errors.append(f"simulator.{field} mismatch")
        elif observed != expected_value:
            errors.append(f"simulator.{field} mismatch")

    runtime = plan.get("runtime", {})
    plan_runtime_map = {
        "matlab_version_full": EXPECTED_RUNTIME["matlab_version_full"],
        "matlab_release": EXPECTED_RUNTIME["matlab_release"],
        "matlab_build": EXPECTED_RUNTIME["matlab_build"],
        "matlab_product_date": EXPECTED_RUNTIME["matlab_product_date"],
        "matlab_runtime_update_date": EXPECTED_RUNTIME["matlab_runtime_update_date"],
        "simulink_version": EXPECTED_RUNTIME["simulink_version"],
        "simulink_release": EXPECTED_RUNTIME["simulink_release"],
        "simulink_product_date": EXPECTED_RUNTIME["simulink_product_date"],
        "architecture": EXPECTED_RUNTIME["architecture"],
        "matlabroot": EXPECTED_RUNTIME["matlabroot"],
    }
    if set(runtime) != set(plan_runtime_map):
        errors.append("runtime field set mismatch")
    for field, expected_value in plan_runtime_map.items():
        if runtime.get(field) != expected_value:
            errors.append(f"runtime.{field} mismatch")

    statistics = plan.get("statistics", {})
    statistics_expected = {
        "independent_unit": "physical_run",
        "agent_case_observations_independent": False,
        "bootstrap": "paired_cluster_bootstrap_stratified_by_true_fault",
        "bootstrap_draws": 10_000,
        "bootstrap_seed": 310031,
        "primary_analysis": "Experiment_3_only",
        "primary_contrast": "B-A",
        "semantic_specificity_contrast": "B-E",
        "experiment_1_plus_experiment_3": "secondary_descriptive_only",
    }
    for field, expected_value in statistics_expected.items():
        if statistics.get(field) != expected_value:
            errors.append(f"statistics.{field} mismatch")
    return errors


def load_attempt_log(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("attempts"), list):
        raise ValueError("attempt log must be an object containing an attempts array")
    return payload


def validate_attempt_log(
    payload: dict[str, Any],
    plan: dict[str, Any],
    generation_script: Path,
    case_plan_path: Path,
    attempt_schema: dict[str, Any],
) -> list[str]:
    errors: list[str] = []
    schema_validator = Draft202012Validator(
        attempt_schema, format_checker=FormatChecker()
    )
    for error in sorted(
        schema_validator.iter_errors(payload),
        key=lambda item: tuple(str(part) for part in item.path),
    ):
        location = ".".join(str(part) for part in error.path) or "<root>"
        errors.append(f"attempt-log schema {location}: {error.message}")
    if payload.get("schema_version") != "1.0":
        errors.append("attempt log schema_version mismatch")
    attempts = payload.get("attempts", [])
    by_case = {row["physical_case_id"]: row for row in plan["cases"]}
    expected_script_hash = sha256_file(generation_script)
    hotfix_002 = json.loads(HOTFIX_002_PATH.read_text(encoding="utf-8"))
    hotfix_002_script_hash = hotfix_002["hotfix_002_generator_sha256"]
    expected_case_plan_hash = sha256_file(case_plan_path)
    seen: set[tuple[str, int]] = set()
    prior_by_case: dict[str, dict[int, dict[str, Any]]] = {}

    for position, row in enumerate(attempts):
        label = f"attempts[{position}]"
        if not isinstance(row, dict):
            errors.append(f"{label} must be an object")
            continue
        missing = ATTEMPT_REQUIRED_FIELDS - row.keys()
        extra = row.keys() - ATTEMPT_REQUIRED_FIELDS
        if missing:
            errors.append(f"{label} missing fields: {sorted(missing)}")
        if extra:
            errors.append(f"{label} unexpected fields: {sorted(extra)}")
        if missing:
            continue

        case_id = row["physical_case_id"]
        attempt = row["attempt"]
        if case_id not in by_case:
            errors.append(f"{label} unknown physical_case_id {case_id!r}")
            continue
        if type(attempt) is not int or attempt not in {0, 1}:
            errors.append(f"{label} attempt must be 0 or 1")
            continue
        key = (case_id, attempt)
        if key in seen:
            errors.append(f"duplicate attempt {key}")
        seen.add(key)

        case = by_case[case_id]
        expected_seed = (
            case["primary_seed"] if attempt == 0 else case["replacement_seed"]
        )
        if row["seed"] != expected_seed:
            errors.append(f"{label} seed mismatch")
        if row["fault_status"] != case["condition"]:
            errors.append(f"{label} fault/status mismatch")
        if row["rng_algorithm"] != "twister":
            errors.append(f"{label} rng_algorithm mismatch")

        for field, expected_value in {**EXPECTED_RUNTIME, **EXPECTED_SIMULATOR}.items():
            if row[field] != expected_value:
                errors.append(f"{label} {field} mismatch")
        legacy_failure_hash_allowed = (
            case_id == "EXP3-N-001"
            and attempt == 0
            and row["seed"] == 310001
            and row["structural_valid"] is False
            and populated(row["technical_failure_reason"])
            and row["generation_script_hash"] == hotfix_002_script_hash
        )
        if (
            row["generation_script_hash"] != expected_script_hash
            and not legacy_failure_hash_allowed
        ):
            errors.append(f"{label} generation_script_hash mismatch")
        if not HASH_PATTERN.fullmatch(row["generation_script_hash"]):
            errors.append(f"{label} invalid generation_script_hash")
        if row["case_plan_hash"] != expected_case_plan_hash:
            errors.append(f"{label} case_plan_hash mismatch")
        if not HASH_PATTERN.fullmatch(row["case_plan_hash"]):
            errors.append(f"{label} invalid case_plan_hash")
        if not populated(row["started_at"]) or not populated(row["completed_at"]):
            errors.append(f"{label} timestamps must be populated")
        expected_name = f"{case_id}__attempt-{attempt}.xlsx"
        if Path(str(row["output_path"])).name != expected_name:
            errors.append(f"{label} output path/case mismatch")
        if row["output_size_bytes"] < 0:
            errors.append(f"{label} output size must be non-negative")
        if row["output_size_bytes"] > 0 and not HASH_PATTERN.fullmatch(
            str(row["output_sha256"])
        ):
            errors.append(f"{label} materialized output SHA-256 missing")
        if row["output_size_bytes"] == 0 and row["output_sha256"] != "":
            errors.append(f"{label} absent output must not have a SHA-256")

        prior = prior_by_case.setdefault(case_id, {})
        if attempt == 1:
            primary = prior.get(0)
            if primary is None:
                errors.append(f"{label} replacement has no earlier attempt 0")
            elif primary["structural_valid"]:
                errors.append(f"{label} replacement follows a valid attempt 0")
            elif not populated(primary["technical_failure_reason"]):
                errors.append(f"{label} attempt 0 lacks technical failure reason")
        prior[attempt] = row

        structural_valid = row["structural_valid"]
        if type(structural_valid) is not bool:
            errors.append(f"{label} structural_valid must be Boolean")
        elif structural_valid:
            if populated(row["technical_failure_reason"]):
                errors.append(f"{label} valid attempt has a failure reason")
            if row["rows"] != 3001 or row["cols"] != 54:
                errors.append(f"{label} valid attempt shape mismatch")
            if row["finite_check"] is not True:
                errors.append(f"{label} valid attempt finite_check must be true")
            if row["output_size_bytes"] <= 0:
                errors.append(f"{label} valid attempt output size must be positive")
            if not HASH_PATTERN.fullmatch(str(row["output_sha256"])):
                errors.append(f"{label} valid attempt output SHA-256 missing")
            if not close(float(row["time_start"]), 0.0):
                errors.append(f"{label} valid attempt Time start mismatch")
            if not close(float(row["time_end"]), 50.0):
                errors.append(f"{label} valid attempt Time end mismatch")
            if not close(float(row["sampling_interval"]), 1 / 60):
                errors.append(f"{label} valid attempt sampling mismatch")
        elif not populated(row["technical_failure_reason"]):
            errors.append(f"{label} invalid attempt lacks technical failure reason")

    for case_id, indexed in prior_by_case.items():
        if len(indexed) > 2 or any(attempt > 1 for attempt in indexed):
            errors.append(f"{case_id} exceeds the attempt limit")
    return errors


def inspect_workbook(path: Path) -> dict[str, Any]:
    try:
        with ZipFile(path) as archive:
            if archive.testzip() is not None:
                raise ValueError("damaged XLSX ZIP container")
    except (BadZipFile, OSError) as exc:
        raise ValueError("invalid XLSX ZIP container") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise ValueError(f"worksheet mismatch: {workbook.sheetnames!r}")
        sheet = workbook["Sheet1"]
        rows = sheet.iter_rows(values_only=True)
        header = list(next(rows))
        if header != EXPECTED_HEADER:
            raise ValueError("header mismatch")
        time_values: list[float] = []
        finite = True
        row_count = 0
        for row_count, row in enumerate(rows, start=1):
            if len(row) != 54:
                raise ValueError(f"row {row_count + 1} has {len(row)} columns")
            for column, value in enumerate(row):
                if isinstance(value, bool) or not isinstance(value, (int, float)):
                    finite = False
                    continue
                number = float(value)
                finite = finite and math.isfinite(number)
                if column == 0:
                    time_values.append(number)
        if len(time_values) < 2:
            raise ValueError("Time has fewer than two samples")
        intervals = [b - a for a, b in zip(time_values, time_values[1:])]
        return {
            "rows": row_count,
            "cols": sheet.max_column,
            "time_start": time_values[0],
            "time_end": time_values[-1],
            "sampling": intervals[0],
            "time_monotonic": all(value > 0 for value in intervals),
            "sampling_constant": all(close(value, 1 / 60) for value in intervals),
            "finite_check": finite,
            "structural_valid": (
                row_count == 3001
                and sheet.max_column == 54
                and close(time_values[0], 0.0)
                and close(time_values[-1], 50.0)
                and all(value > 0 for value in intervals)
                and all(close(value, 1 / 60) for value in intervals)
                and finite
            ),
        }
    finally:
        workbook.close()


def load_manifest(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as stream:
        reader = csv.DictReader(stream)
        if tuple(reader.fieldnames or ()) != MANIFEST_FIELDS:
            raise ValueError(
                "manifest columns do not exactly match the frozen template"
            )
        return list(reader)


def validate_manifest_template(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    expected = canonical_cases()
    if len(rows) != 30:
        return [f"manifest template must contain 30 rows, found {len(rows)}"]
    for row, case in zip(rows, expected):
        if row["physical_case_id"] != case["physical_case_id"]:
            errors.append("manifest template case order/ID mismatch")
        if row["fault/status"] != case["condition"]:
            errors.append(f"{case['physical_case_id']} condition mismatch")
        if row["seed"] != str(case["primary_seed"]):
            errors.append(f"{case['physical_case_id']} primary seed mismatch")
        for field in MANIFEST_FIELDS:
            if field not in {"physical_case_id", "fault/status", "seed"}:
                if row[field] != "TBD":
                    errors.append(f"{case['physical_case_id']} {field} must be TBD")
    return errors


def validate_manifest(
    rows: list[dict[str, str]],
    plan: dict[str, Any],
    attempt_log: dict[str, Any],
    data_dir: Path,
) -> list[str]:
    errors: list[str] = []
    by_case = {row["physical_case_id"]: row for row in plan["cases"]}
    attempts = {
        (row["physical_case_id"], row["attempt"]): row
        for row in attempt_log["attempts"]
        if isinstance(row, dict)
        and row.get("physical_case_id") in by_case
        and row.get("attempt") in {0, 1}
    }
    ids = [row["physical_case_id"] for row in rows]
    if len(rows) != 30 or set(ids) != set(by_case) or len(set(ids)) != 30:
        errors.append("manifest must contain exactly the 30 unique intended cases")
        return errors

    filenames: set[str] = set()
    for row in rows:
        case_id = row["physical_case_id"]
        case = by_case[case_id]
        label = f"manifest[{case_id}]"
        try:
            attempt = int(row["attempt"])
            seed = int(row["seed"])
            size = int(row["size_bytes"])
            expected_rows = int(row["rows"])
            expected_cols = int(row["cols"])
            expected_start = float(row["time_start"])
            expected_end = float(row["time_end"])
            expected_sampling = float(row["sampling"])
            finite_check = parse_bool(row["finite_check"], "finite_check")
            structural_valid = parse_bool(row["structural_valid"], "structural_valid")
        except (TypeError, ValueError) as exc:
            errors.append(f"{label} invalid typed field: {exc}")
            continue

        if attempt not in {0, 1}:
            errors.append(f"{label} attempt must be 0 or 1")
            continue
        expected_seed = (
            case["primary_seed"] if attempt == 0 else case["replacement_seed"]
        )
        if row["fault/status"] != case["condition"] or seed != expected_seed:
            errors.append(f"{label} case/condition/seed mapping mismatch")
        expected_filename = f"{case_id}__attempt-{attempt}.xlsx"
        if (
            row["filename"] != expected_filename
            or Path(row["filename"]).name != row["filename"]
        ):
            errors.append(f"{label} filename/case mismatch")
        if row["filename"] in filenames:
            errors.append(f"{label} duplicate filename")
        filenames.add(row["filename"])
        if not HASH_PATTERN.fullmatch(row["SHA256"]):
            errors.append(f"{label} invalid SHA-256")
        if not finite_check or not structural_valid:
            errors.append(f"{label} accepted manifest row must be structurally valid")

        attempt_record = attempts.get((case_id, attempt))
        if attempt_record is None or not attempt_record.get("structural_valid"):
            errors.append(f"{label} lacks matching valid attempt-log record")
        else:
            comparisons = {
                "output_size_bytes": size,
                "output_sha256": row["SHA256"],
                "rows": expected_rows,
                "cols": expected_cols,
                "finite_check": finite_check,
                "structural_valid": structural_valid,
            }
            for field, expected_value in comparisons.items():
                if attempt_record[field] != expected_value:
                    errors.append(f"{label} {field} disagrees with attempt log")
            for field, expected_value in {
                "time_start": expected_start,
                "time_end": expected_end,
                "sampling_interval": expected_sampling,
            }.items():
                if not close(float(attempt_record[field]), expected_value):
                    errors.append(f"{label} {field} disagrees with attempt log")

        path = data_dir / row["filename"]
        if not path.is_file():
            errors.append(f"{label} missing workbook")
            continue
        if path.stat().st_size != size:
            errors.append(f"{label} size mismatch")
        if sha256_file(path) != row["SHA256"]:
            errors.append(f"{label} SHA-256 mismatch")
        try:
            observed = inspect_workbook(path)
        except Exception as exc:
            errors.append(f"{label} workbook error: {exc}")
            continue
        expected_observed = {
            "rows": expected_rows,
            "cols": expected_cols,
            "finite_check": finite_check,
            "structural_valid": structural_valid,
        }
        for field, expected_value in expected_observed.items():
            if observed[field] != expected_value:
                errors.append(f"{label} observed {field} mismatch")
        for field, expected_value in {
            "time_start": expected_start,
            "time_end": expected_end,
            "sampling": expected_sampling,
        }.items():
            if not close(float(observed[field]), expected_value):
                errors.append(f"{label} observed {field} mismatch")

    logged_files: set[str] = set()
    for attempt_record in attempt_log["attempts"]:
        if not isinstance(attempt_record, dict):
            continue
        try:
            size = int(attempt_record["output_size_bytes"])
        except (KeyError, TypeError, ValueError):
            continue
        if size <= 0:
            continue
        logged_path = Path(str(attempt_record["output_path"]))
        logged_files.add(logged_path.name)
        materialized_path = data_dir / logged_path.name
        if not materialized_path.is_file():
            errors.append(f"logged materialized output missing: {logged_path.name}")
            continue
        if materialized_path.stat().st_size != size:
            errors.append(
                f"logged materialized output size mismatch: {logged_path.name}"
            )
        if sha256_file(materialized_path) != attempt_record["output_sha256"]:
            errors.append(
                f"logged materialized output hash mismatch: {logged_path.name}"
            )

    observed_files = {path.name for path in data_dir.glob("*.xlsx")}
    if observed_files != logged_files:
        missing = logged_files - observed_files
        extra = observed_files - logged_files
        errors.append(
            f"data directory/attempt-log mismatch: missing={sorted(missing)}, "
            f"extra={sorted(extra)}"
        )
    return errors


def prefreeze_checks(
    case_plan_path: Path,
    manifest_template_path: Path,
    attempt_template_path: Path,
    attempt_schema_path: Path,
    freeze_manifest_path: Path,
) -> list[str]:
    errors: list[str] = []
    try:
        plan = load_case_plan(case_plan_path)
    except Exception as exc:
        return [str(exc)]
    try:
        template_rows = load_manifest(manifest_template_path)
        errors.extend(validate_manifest_template(template_rows))
    except Exception as exc:
        errors.append(f"manifest template: {exc}")
    try:
        attempt_template = load_attempt_log(attempt_template_path)
        attempt_schema = json.loads(attempt_schema_path.read_text(encoding="utf-8"))
        Draft202012Validator.check_schema(attempt_schema)
        schema_errors = Draft202012Validator(
            attempt_schema, format_checker=FormatChecker()
        ).iter_errors(attempt_template)
        errors.extend(
            f"attempt-log template schema: {error.message}" for error in schema_errors
        )
        if attempt_template.get("schema_version") != "1.0":
            errors.append("attempt-log template schema_version mismatch")
        if attempt_template.get("attempts") != []:
            errors.append("attempt-log template must contain zero attempts")
    except Exception as exc:
        errors.append(f"attempt-log template: {exc}")
    if plan["rng"]["max_total_attempts"] != 2:
        errors.append("replacement policy does not allow exactly two total attempts")
    if plan.get("status") != "FROZEN_BEFORE_GENERATION":
        errors.append("case plan is not FROZEN_BEFORE_GENERATION")
    errors.extend(validate_freeze_manifest(freeze_manifest_path))
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--case-plan", type=Path, default=SCRIPT_DIR / "exp3_case_plan.json"
    )
    parser.add_argument(
        "--generation-script",
        type=Path,
        default=SCRIPT_DIR / "generate_exp3_heldout.m",
    )
    parser.add_argument(
        "--freeze-manifest",
        type=Path,
        default=SCRIPT_DIR / "EXP3_FREEZE_MANIFEST.json",
    )
    parser.add_argument(
        "--attempt-log",
        type=Path,
        default=ROOT / "tep_exp3_heldout/exp3_attempt_log.json",
    )
    parser.add_argument(
        "--manifest", type=Path, default=ROOT / "tep_exp3_heldout/exp3_manifest.csv"
    )
    parser.add_argument(
        "--data-dir", type=Path, default=ROOT / "tep_exp3_heldout/mode1"
    )
    parser.add_argument("--pre-freeze", action="store_true")
    args = parser.parse_args()

    try:
        if args.pre_freeze:
            errors = prefreeze_checks(
                args.case_plan,
                SCRIPT_DIR / "exp3_manifest_template.csv",
                SCRIPT_DIR / "exp3_attempt_log.template.json",
                SCRIPT_DIR / "exp3_attempt_log.schema.json",
                args.freeze_manifest,
            )
        else:
            plan = load_case_plan(args.case_plan)
            if plan.get("status") != "FROZEN_BEFORE_GENERATION":
                raise ValueError(
                    "case plan is not frozen; held-out verification is disabled"
                )
            freeze_errors = validate_freeze_manifest(args.freeze_manifest)
            if freeze_errors:
                raise ValueError("; ".join(freeze_errors))
            attempt_log = load_attempt_log(args.attempt_log)
            attempt_schema = json.loads(
                (SCRIPT_DIR / "exp3_attempt_log.schema.json").read_text(
                    encoding="utf-8"
                )
            )
            errors = validate_attempt_log(
                attempt_log,
                plan,
                args.generation_script,
                args.case_plan,
                attempt_schema,
            )
            manifest = load_manifest(args.manifest)
            errors.extend(validate_manifest(manifest, plan, attempt_log, args.data_dir))
    except Exception as exc:
        print(f"FAIL: {exc}", file=sys.stderr)
        return 1

    if errors:
        print("FAIL: Experiment 3 verification failed", file=sys.stderr)
        for error in errors:
            print(f"  - {error}", file=sys.stderr)
        return 1
    mode = "pre-freeze infrastructure" if args.pre_freeze else "held-out"
    print(f"PASS: Experiment 3 {mode} verification succeeded.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
