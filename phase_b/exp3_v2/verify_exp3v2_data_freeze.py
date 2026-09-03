#!/usr/bin/env python3
"""Portable verifier for the disconnected EXP3_V2 data-freeze history."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path, PurePosixPath
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


MANIFEST_RELATIVE = Path("phase_b/exp3_v2/EXP3_V2_DATA_FREEZE_MANIFEST_001.json")
VERIFIER_RELATIVE = Path("phase_b/exp3_v2/verify_exp3v2_data_freeze.py")
ATTEMPT_LOG_RELATIVE = Path("tep_exp3_v2_heldout/exp3v2_attempt_log.json")
POPULATED_MANIFEST_RELATIVE = Path("tep_exp3_v2_heldout/exp3v2_manifest.csv")
DATA_RELATIVE = Path("tep_exp3_v2_heldout/mode1")
EXPECTED_EXPERIMENT = "Experiment 3 V2 — Prospective Fresh-Run Held-Out"
EXPECTED_TAG = "exp3-v2-heldout-data-frozen-001"
EXPECTED_CONDITIONS = ("Normal", "F1", "F8", "F10", "F13")
EXPECTED_IDS = tuple(
    [f"EXP3V2-N-{index:03d}" for index in range(1, 7)]
    + [f"EXP3V2-F1-{index:03d}" for index in range(1, 7)]
    + [f"EXP3V2-F8-{index:03d}" for index in range(1, 7)]
    + [f"EXP3V2-F10-{index:03d}" for index in range(1, 7)]
    + [f"EXP3V2-F13-{index:03d}" for index in range(1, 7)]
)
EXPECTED_HEADER = tuple(
    ["Time (h)"]
    + [f"XMEAS-{index}" for index in range(1, 42)]
    + [f"XMV-{index}" for index in range(1, 13)]
)
MANIFEST_FIELDS = (
    "physical_case_id",
    "fault/status",
    "attempt",
    "seed",
    "filename",
    "size_bytes",
    "SHA256",
    "rows",
    "cols",
    "time_start",
    "time_end",
    "sampling",
    "finite_check",
    "structural_valid",
)
EXPECTED_RUNTIME = {
    "matlab_version_full": "25.2.0.3312555 (R2025b) Update 6",
    "matlab_release": "2025b",
    "matlab_build": "3312555",
    "matlab_product_date": "28-Jul-2025",
    "matlab_runtime_update_date": "June 30, 2026",
    "simulink_version": "25.2",
    "simulink_release": "(R2025b)",
    "simulink_product_date": "28-Jul-2025",
    "architecture": "MACA64",
    "matlabroot": "/Applications/MATLAB_R2025b.app",
}
EXPECTED_SIMULATOR = {
    "rng_algorithm": "twister",
    "simulator_commit": "a0413e16c940f0fc8b554d6a86248020d7fb7527",
    "model_name": "MultiLoop_mode1",
    "simulation_mode": "normal",
    "solver": "ode45",
    "sfunction_identity": "temexd_mod",
    "sfunction_hash": (
        "0da41d939e5ab7ba122d7b70c124368ee0882fce40e775dba5d180e7a7e24e5e"
    ),
    "sfunction_mex_hash": (
        "68f632388cb698dd7b8c595000bc03c2e1d19200546b9d4357df90e3fc93af0d"
    ),
    "model_hash": ("d2f6659f65935021d4b1813e7189be02e7ae9f5639b794e8edc4f2f3c5cddba8"),
    "initial_state_hash": (
        "40eaebc92badb04ad026e358cfd28ec9c778fcf2d24a1b8f5d85565854da2747"
    ),
    "generation_script_hash": (
        "4e746a8b6504953d2bb0d4eb9982cdef1ee3c02d0d0f1cb374e5a9086e45a9f1"
    ),
}
ATTEMPT_FIELDS = {
    "physical_case_id",
    "fault_status",
    "attempt",
    "seed",
    "rng_algorithm",
    "started_at",
    "completed_at",
    *EXPECTED_RUNTIME,
    "simulator_commit",
    "model_name",
    "simulation_mode",
    "solver",
    "sfunction_identity",
    "sfunction_hash",
    "sfunction_mex_hash",
    "model_hash",
    "initial_state_hash",
    "case_plan_hash",
    "generation_script_hash",
    "output_path",
    "output_size_bytes",
    "output_sha256",
    "rows",
    "cols",
    "time_start",
    "time_end",
    "sampling_interval",
    "finite_check",
    "structural_valid",
    "technical_failure_reason",
}
TIME_TOLERANCE = 1e-10


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def close(left: float, right: float) -> bool:
    return math.isclose(left, right, rel_tol=0.0, abs_tol=TIME_TOLERANCE)


def is_json_integer(value: Any) -> bool:
    return (
        not isinstance(value, bool)
        and isinstance(value, (int, float))
        and math.isfinite(float(value))
        and float(value).is_integer()
    )


def canonical_condition(case_id: str) -> str:
    token = case_id.split("-")[1]
    return "Normal" if token == "N" else token


def parse_boolean(value: str, label: str) -> bool:
    if value not in {"true", "false"}:
        raise ValueError(f"{label} must be lowercase true/false")
    return value == "true"


def load_json(path: Path, label: str) -> dict[str, Any]:
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise ValueError(f"{label} is not valid UTF-8 JSON: {exc}") from exc
    if not isinstance(payload, dict):
        raise ValueError(f"{label} must be a JSON object")
    return payload


def validate_attempt_schema(payload: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if set(payload) != {"schema_version", "experiment", "attempts"}:
        errors.append("attempt log top-level schema mismatch")
        return errors
    if payload.get("schema_version") != "1.0":
        errors.append("attempt log schema_version mismatch")
    if payload.get("experiment") != EXPECTED_EXPERIMENT:
        errors.append("attempt log experiment mismatch")
    attempts = payload.get("attempts")
    if not isinstance(attempts, list):
        return errors + ["attempt log attempts must be an array"]
    if len(attempts) != 30:
        errors.append("attempt log must contain exactly 30 records")
    for position, row in enumerate(attempts):
        label = f"attempts[{position}]"
        if not isinstance(row, dict) or set(row) != ATTEMPT_FIELDS:
            errors.append(f"{label} field-set mismatch")
            continue
        expected_id = EXPECTED_IDS[position] if position < len(EXPECTED_IDS) else None
        if row["physical_case_id"] != expected_id:
            errors.append(f"{label} canonical case-order mismatch")
        if row["fault_status"] != canonical_condition(str(expected_id)):
            errors.append(f"{label} condition mismatch")
        if not is_json_integer(row["attempt"]) or row["attempt"] != 0:
            errors.append(f"{label} must be attempt 0")
        if not is_json_integer(row["seed"]) or row["seed"] != 320001 + position:
            errors.append(f"{label} seed mismatch")
        for field, expected in {**EXPECTED_RUNTIME, **EXPECTED_SIMULATOR}.items():
            if row[field] != expected:
                errors.append(f"{label} {field} mismatch")
        for field in ("started_at", "completed_at"):
            if not isinstance(row[field], str):
                errors.append(f"{label} {field} must be a string")
                continue
            try:
                datetime.fromisoformat(row[field].replace("Z", "+00:00"))
            except ValueError:
                errors.append(f"{label} {field} is not an ISO date-time")
        if (
            not isinstance(row["output_path"], str)
            or not Path(row["output_path"]).is_absolute()
        ):
            errors.append(f"{label} output_path must retain absolute provenance")
        expected_name = f"{expected_id}__attempt-0.xlsx"
        if PurePosixPath(str(row["output_path"])).name != expected_name:
            errors.append(f"{label} output_path basename mismatch")
        integer_fields = ("output_size_bytes", "rows", "cols")
        if any(not is_json_integer(row[field]) for field in integer_fields):
            errors.append(f"{label} integer-field type mismatch")
        number_fields = ("time_start", "time_end", "sampling_interval")
        if any(
            isinstance(row[field], bool) or not isinstance(row[field], (int, float))
            for field in number_fields
        ):
            errors.append(f"{label} numeric-field type mismatch")
        if type(row["finite_check"]) is not bool or row["finite_check"] is not True:
            errors.append(f"{label} finite_check mismatch")
        if (
            type(row["structural_valid"]) is not bool
            or row["structural_valid"] is not True
        ):
            errors.append(f"{label} structural_valid mismatch")
        if row["technical_failure_reason"] != "":
            errors.append(f"{label} technical_failure_reason must be empty")
        if (
            row["rows"] != 3001
            or row["cols"] != 54
            or not close(float(row["time_start"]), 0.0)
            or not close(float(row["time_end"]), 50.0)
            or not close(float(row["sampling_interval"]), 1 / 60)
        ):
            errors.append(f"{label} structural metadata mismatch")
    return errors


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as stream:
        reader = csv.DictReader(stream)
        if tuple(reader.fieldnames or ()) != MANIFEST_FIELDS:
            raise ValueError("populated manifest field-set mismatch")
        return list(reader)


def inspect_workbook(path: Path) -> dict[str, Any]:
    try:
        with ZipFile(path) as archive:
            if archive.testzip() is not None:
                raise ValueError("damaged XLSX ZIP container")
    except (BadZipFile, OSError) as exc:
        raise ValueError("invalid XLSX ZIP container") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise ValueError("worksheet mismatch")
        rows = workbook["Sheet1"].iter_rows(values_only=True)
        if tuple(next(rows)) != EXPECTED_HEADER:
            raise ValueError("header mismatch")
        times: list[float] = []
        finite = True
        count = 0
        for count, row in enumerate(rows, start=1):
            if len(row) != 54:
                raise ValueError("column count mismatch")
            for column, value in enumerate(row):
                numeric = not isinstance(value, bool) and isinstance(
                    value, (int, float)
                )
                finite = finite and numeric and math.isfinite(float(value))
                if column == 0 and numeric:
                    times.append(float(value))
        if len(times) != count or not times:
            raise ValueError("time-axis numeric coverage mismatch")
        intervals = [right - left for left, right in zip(times, times[1:])]
        return {
            "rows": count,
            "cols": workbook["Sheet1"].max_column,
            "time_start": times[0],
            "time_end": times[-1],
            "sampling": intervals[0],
            "finite_check": finite,
            "structural_valid": count == 3001
            and workbook["Sheet1"].max_column == 54
            and close(times[0], 0.0)
            and close(times[-1], 50.0)
            and len(intervals) == 3000
            and all(value > 0 and close(value, 1 / 60) for value in intervals)
            and finite,
        }
    finally:
        workbook.close()


def validate(root: Path, freeze_manifest_path: Path) -> list[str]:
    errors: list[str] = []
    freeze = load_json(freeze_manifest_path, "data-freeze manifest")
    if freeze.get("status") != "FROZEN_BEFORE_VERBALIZATION":
        errors.append("data-freeze status mismatch")
    if freeze.get("prospective_tag") != EXPECTED_TAG:
        errors.append("prospective tag mismatch")
    portable = freeze.get("portable_data_verifier", {})
    if portable.get("path") != VERIFIER_RELATIVE.as_posix():
        errors.append("portable verifier path binding mismatch")
    if portable.get("sha256") != sha256_file(root / VERIFIER_RELATIVE):
        errors.append("portable verifier hash mismatch")

    data_artifacts = freeze.get("data_artifacts", {})
    log_contract = data_artifacts.get("attempt_log", {})
    csv_contract = data_artifacts.get("populated_manifest", {})
    log_path = root / ATTEMPT_LOG_RELATIVE
    csv_path = root / POPULATED_MANIFEST_RELATIVE
    if log_contract.get("path") != ATTEMPT_LOG_RELATIVE.as_posix():
        errors.append("attempt-log relative path binding mismatch")
    if csv_contract.get("path") != POPULATED_MANIFEST_RELATIVE.as_posix():
        errors.append("populated-manifest relative path binding mismatch")
    for path, contract, label in (
        (log_path, log_contract, "attempt log"),
        (csv_path, csv_contract, "populated manifest"),
    ):
        if path.is_symlink() or not path.is_file():
            errors.append(f"{label} missing or symlinked")
        elif path.stat().st_size != contract.get("size_bytes"):
            errors.append(f"{label} size mismatch")
        elif sha256_file(path) != contract.get("sha256"):
            errors.append(f"{label} hash mismatch")
    if errors:
        return errors

    attempt_log = load_json(log_path, "attempt log")
    errors.extend(validate_attempt_schema(attempt_log))
    rows = load_csv(csv_path)
    if len(rows) != 30:
        return errors + ["populated manifest must contain exactly 30 rows"]
    inventory = data_artifacts.get("workbook_inventory")
    if not isinstance(inventory, list) or len(inventory) != 30:
        return errors + ["data-freeze workbook inventory must contain 30 rows"]
    attempts = attempt_log["attempts"]
    inventory_digest = hashlib.sha256()
    concatenated_digest = hashlib.sha256()
    expected_workbook_paths: set[str] = set()
    for position, (row, attempt, item) in enumerate(zip(rows, attempts, inventory), 1):
        label = f"case[{position}]"
        expected_id = EXPECTED_IDS[position - 1]
        expected_condition = canonical_condition(expected_id)
        expected_name = f"{expected_id}__attempt-0.xlsx"
        expected_relative = (DATA_RELATIVE / expected_name).as_posix()
        expected_workbook_paths.add(expected_relative)
        try:
            typed = {
                "attempt": int(row["attempt"]),
                "seed": int(row["seed"]),
                "size_bytes": int(row["size_bytes"]),
                "rows": int(row["rows"]),
                "cols": int(row["cols"]),
                "time_start": float(row["time_start"]),
                "time_end": float(row["time_end"]),
                "sampling": float(row["sampling"]),
                "finite_check": parse_boolean(row["finite_check"], label),
                "structural_valid": parse_boolean(row["structural_valid"], label),
            }
        except (TypeError, ValueError) as exc:
            errors.append(f"{label} typed manifest value invalid: {exc}")
            continue
        if (
            row["physical_case_id"] != expected_id
            or row["fault/status"] != expected_condition
            or row["filename"] != expected_name
            or typed["attempt"] != 0
            or typed["seed"] != 320000 + position
        ):
            errors.append(f"{label} manifest identity/order mismatch")
        if (
            item.get("order") != position
            or item.get("physical_case_id") != expected_id
            or item.get("condition") != expected_condition
            or item.get("attempt") != 0
            or item.get("seed") != 320000 + position
            or item.get("filename") != expected_name
            or item.get("path") != expected_relative
        ):
            errors.append(f"{label} data-freeze inventory identity mismatch")
        if PurePosixPath(str(attempt["output_path"])).name != expected_name:
            errors.append(f"{label} immutable output_path basename mismatch")
        comparisons = {
            "output_size_bytes": typed["size_bytes"],
            "output_sha256": row["SHA256"],
            "rows": typed["rows"],
            "cols": typed["cols"],
        }
        for field, expected in comparisons.items():
            if attempt[field] != expected:
                errors.append(f"{label} attempt-log {field} mismatch")
        if item.get("size_bytes") != typed["size_bytes"]:
            errors.append(f"{label} inventory size mismatch")
        if item.get("sha256") != row["SHA256"]:
            errors.append(f"{label} inventory hash mismatch")
        workbook_path = root / expected_relative
        if workbook_path.is_symlink() or not workbook_path.is_file():
            errors.append(f"{label} workbook missing or symlinked")
            continue
        raw_hash = hashlib.sha256()
        with workbook_path.open("rb") as stream:
            for block in iter(lambda: stream.read(1024 * 1024), b""):
                raw_hash.update(block)
                concatenated_digest.update(block)
        actual_hash = raw_hash.hexdigest()
        actual_size = workbook_path.stat().st_size
        if actual_size != typed["size_bytes"] or actual_hash != row["SHA256"]:
            errors.append(f"{label} workbook size/hash mismatch")
        inventory_digest.update(
            f"{expected_name},{actual_size},{actual_hash}\n".encode("utf-8")
        )
        try:
            observed = inspect_workbook(workbook_path)
        except Exception as exc:
            errors.append(f"{label} workbook structural error: {exc}")
            continue
        expected_structure = {
            "rows": typed["rows"],
            "cols": typed["cols"],
            "time_start": typed["time_start"],
            "time_end": typed["time_end"],
            "sampling": typed["sampling"],
            "finite_check": typed["finite_check"],
            "structural_valid": typed["structural_valid"],
        }
        for field, expected in expected_structure.items():
            actual = observed[field]
            if isinstance(expected, float):
                if not close(float(actual), expected):
                    errors.append(f"{label} observed {field} mismatch")
            elif actual != expected:
                errors.append(f"{label} observed {field} mismatch")

    observed_workbooks = {
        path.relative_to(root).as_posix()
        for path in (root / DATA_RELATIVE).glob("*.xlsx")
    }
    if observed_workbooks != expected_workbook_paths:
        errors.append("workbook path set mismatch")
    if Counter(row["fault/status"] for row in rows) != Counter(
        {condition: 6 for condition in EXPECTED_CONDITIONS}
    ):
        errors.append("condition counts mismatch")
    aggregate = data_artifacts.get("aggregate_digests", {})
    if aggregate.get("workbook_count") != 30:
        errors.append("aggregate workbook count mismatch")
    if aggregate.get("total_workbook_bytes") != sum(
        int(row["size_bytes"]) for row in rows
    ):
        errors.append("aggregate workbook byte count mismatch")
    if aggregate.get("inventory_sha256") != inventory_digest.hexdigest():
        errors.append("aggregate inventory digest mismatch")
    if (
        aggregate.get("concatenated_workbook_bytes_sha256")
        != concatenated_digest.hexdigest()
    ):
        errors.append("aggregate concatenated-byte digest mismatch")

    expected_tree = {
        ATTEMPT_LOG_RELATIVE.as_posix(),
        POPULATED_MANIFEST_RELATIVE.as_posix(),
        MANIFEST_RELATIVE.as_posix(),
        VERIFIER_RELATIVE.as_posix(),
        *expected_workbook_paths,
    }
    observed_tree = {
        path.relative_to(root).as_posix()
        for path in root.rglob("*")
        if path.is_file() and ".git" not in path.relative_to(root).parts
    }
    if observed_tree != expected_tree:
        errors.append(
            "checked-out tree mismatch: "
            f"missing={sorted(expected_tree - observed_tree)}, "
            f"extra={sorted(observed_tree - expected_tree)}"
        )
    return errors


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--root",
        type=Path,
        default=Path(__file__).resolve().parents[2],
        help="root of the checked-out disconnected data-freeze tree",
    )
    parser.add_argument(
        "--freeze-manifest",
        type=Path,
        help="override the data-freeze manifest path",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    root = args.root.resolve()
    manifest = (
        args.freeze_manifest.resolve()
        if args.freeze_manifest
        else root / MANIFEST_RELATIVE
    )
    try:
        errors = validate(root, manifest)
    except Exception as exc:
        errors = [str(exc)]
    if errors:
        for error in errors:
            print(f"FAIL: {error}", file=sys.stderr)
        return 1
    print("PASS: Experiment 3 V2 portable data-freeze verification succeeded.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
