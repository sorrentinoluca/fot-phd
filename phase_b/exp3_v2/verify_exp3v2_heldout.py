#!/usr/bin/env python3
"""Fail-closed structural and provenance verifier for Experiment 3 V2."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import re
import subprocess
import sys
from collections import Counter
from pathlib import Path
from typing import Any
from zipfile import BadZipFile, ZipFile

from jsonschema import Draft202012Validator, FormatChecker
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
REAL_ROOT = (ROOT / "tep_exp3_v2_heldout").resolve()
OLD_REAL_ROOT = (ROOT / "tep_exp3_heldout").resolve()
EXPECTED_CONDITIONS = ("Normal", "F1", "F8", "F10", "F13")
EXPECTED_HEADER = (
    ["Time (h)"]
    + [f"XMEAS-{index}" for index in range(1, 42)]
    + [f"XMV-{index}" for index in range(1, 13)]
)
EXPECTED_RUNTIME = {
    "matlab_version_full": "25.2.0.3312555 (R2025b) Update 6",
    "matlab_release": "2025b",
    "matlab_build": "3312555",
    "matlab_product_date": "28-Jul-2025",
    "matlab_runtime_update_date": "June 30, 2026",
    "simulink_version": "25.2",
    "simulink_release": "(R2025b)",
    "simulink_product_date": "28-Jul-2025",
    "architecture": "MACA64",
    "matlabroot": "/Applications/MATLAB_R2025b.app",
}
EXPECTED_SIMULATOR = {
    "simulator_commit": "a0413e16c940f0fc8b554d6a86248020d7fb7527",
    "model": "MultiLoop_mode1",
    "tep_mode": "Mode 1",
    "simulation_mode": "normal",
    "solver": "ode45",
    "start_time_h": 0.0,
    "stop_time_h": 50.0,
    "sampling_interval_h": 1 / 60,
    "fault_injection_h": 10.0,
    "initial_state_file": "Mode1xInitial.mat",
    "initial_state_sha256": (
        "40eaebc92badb04ad026e358cfd28ec9c778fcf2d24a1b8f5d85565854da2747"
    ),
    "model_sha256": (
        "d2f6659f65935021d4b1813e7189be02e7ae9f5639b794e8edc4f2f3c5cddba8"
    ),
    "sfunction_identity": "temexd_mod",
    "sfunction_source_sha256": (
        "0da41d939e5ab7ba122d7b70c124368ee0882fce40e775dba5d180e7a7e24e5e"
    ),
    "sfunction_macos_mex_sha256": (
        "68f632388cb698dd7b8c595000bc03c2e1d19200546b9d4357df90e3fc93af0d"
    ),
    "custom_setpoints": False,
    "expected_rows": 3001,
    "expected_columns": 54,
    "worksheet": "Sheet1",
}
EXPECTED_ATTEMPT_SIMULATOR = {
    "simulator_commit": EXPECTED_SIMULATOR["simulator_commit"],
    "model_name": EXPECTED_SIMULATOR["model"],
    "simulation_mode": EXPECTED_SIMULATOR["simulation_mode"],
    "solver": EXPECTED_SIMULATOR["solver"],
    "sfunction_identity": EXPECTED_SIMULATOR["sfunction_identity"],
    "sfunction_hash": EXPECTED_SIMULATOR["sfunction_source_sha256"],
    "sfunction_mex_hash": EXPECTED_SIMULATOR["sfunction_macos_mex_sha256"],
    "model_hash": EXPECTED_SIMULATOR["model_sha256"],
    "initial_state_hash": EXPECTED_SIMULATOR["initial_state_sha256"],
}
EXPECTED_EXTERNAL_DEPENDENCIES = {
    "Mode1xInitial.mat": (
        13592,
        "40eaebc92badb04ad026e358cfd28ec9c778fcf2d24a1b8f5d85565854da2747",
    ),
    "Mode_1_Init.m": (
        2270,
        "9dfb4e404c8c982c035fe47472020443b0a1d3f37b55425219968489d92d8933",
    ),
    "MultiLoop_mode1.mdl": (
        186660,
        "d2f6659f65935021d4b1813e7189be02e7ae9f5639b794e8edc4f2f3c5cddba8",
    ),
    "TElib.mdl": (
        12702,
        "4605de6ca0e6da67626e2be6d5f328c735f8bf5a5a730dc67f558a3f1dabddba",
    ),
    "TEplot.m": (
        4871,
        "f10cc8751c1dd99c2efe989460871e701704bc8bde901d83a13834327e75b1be",
    ),
    "temexd_mod.c": (
        195118,
        "0da41d939e5ab7ba122d7b70c124368ee0882fce40e775dba5d180e7a7e24e5e",
    ),
    "temexd_mod.mexmaca64": (
        90232,
        "68f632388cb698dd7b8c595000bc03c2e1d19200546b9d4357df90e3fc93af0d",
    ),
    "teprob_mod.h": (
        6907,
        "e8d07857030a837443ce947361335f2e6f2ade5d2fa54a85bcc5c4a6d9afe939",
    ),
}
EXPECTED_HISTORY = (
    (
        "b02e93f92bf6fa85a4fd0a2e010bac365a3a7c89",
        "exp3-heldout-frozen",
    ),
    (
        "cdba0202435d1c97ea79cfff586e59534ce9baad",
        "exp3-heldout-frozen-hotfix-001",
    ),
    (
        "28130023a34eda778c04a001a9f631404bd6b9a6",
        "exp3-post-freeze-hotfix-002",
    ),
    (
        "0d869720e6ac4d1b396b3b9d731463324d296e26",
        "exp3-post-freeze-hotfix-003",
    ),
    (
        "1cad481839475afaa6ad784bba25c1c45bb260ed",
        "exp3-post-freeze-hotfix-004",
    ),
)
ATTEMPT_LOG_HASH = "04ea7d8af227c3a7f947b4dde434e77510c163ce9c108892ffa22f491f022904"
HASH_RE = re.compile(r"^[0-9a-f]{64}$")
TIME_TOLERANCE = 1e-10
ACTIVE_SENTINEL_ID = "EXP3V2-SENTINEL-002"
ACTIVE_SENTINEL_SEED = 987654322
CONSUMED_SENTINELS = {"EXP3V2-SENTINEL-001": 987654321}
FINAL_CANDIDATE_STATUS = "PENDING_HUMAN_FINAL_FREEZE"
FINAL_REVISION_DRAFT_STATUS = "PRE_FREEZE_DRAFT"
ORIGINAL_FINAL_MANIFEST_HASH = (
    "cbefaefc585d68b66351961bdeb8289cec48079a5964f8bc660c82c5ec95dc5d"
)
SENTINEL_EVIDENCE_HASHES = {
    "EXP3_V2_SENTINEL_EVIDENCE.json": (
        "daf67273138bf192d77e62dd56bc8598a90070baaf4f8714a8851ed3ca9f3a86"
    ),
    "EXP3_V2_SENTINEL_EVIDENCE.md": (
        "84823762a9c1109b565afe0c319999a89eb033016ab523d33896317129ceb227"
    ),
}
MANIFEST_FIELDS = (
    "physical_case_id",
    "fault/status",
    "attempt",
    "seed",
    "filename",
    "size_bytes",
    "SHA256",
    "rows",
    "cols",
    "time_start",
    "time_end",
    "sampling",
    "finite_check",
    "structural_valid",
)
ATTEMPT_FIELDS = {
    "physical_case_id",
    "fault_status",
    "attempt",
    "seed",
    "rng_algorithm",
    "started_at",
    "completed_at",
    *EXPECTED_RUNTIME,
    "simulator_commit",
    "model_name",
    "simulation_mode",
    "solver",
    "sfunction_identity",
    "sfunction_hash",
    "sfunction_mex_hash",
    "model_hash",
    "initial_state_hash",
    "case_plan_hash",
    "generation_script_hash",
    "output_path",
    "output_size_bytes",
    "output_sha256",
    "rows",
    "cols",
    "time_start",
    "time_end",
    "sampling_interval",
    "finite_check",
    "structural_valid",
    "technical_failure_reason",
}
REQUIRED_HARNESS_PATHS = {
    ".gitignore",
    "requirements.txt",
    "phase_b/exp3/EXP3_CLOSURE.md",
    "phase_b/exp3/EXP3_CLOSURE.json",
    "phase_b/exp3/EXP3_CLOSURE_attempt_log_archive.json",
    "phase_b/exp3_v2/EXP3_V2_FRESH_RUN_PROTOCOL.md",
    "phase_b/exp3_v2/EXP3_V2_HARNESS_FREEZE_MANIFEST.json",
    "phase_b/exp3_v2/EXP3_V2_HARNESS_FREEZE_MANIFEST_002.json",
    "phase_b/exp3_v2/EXP3_V2_HARNESS_FREEZE_MANIFEST_003.json",
    "phase_b/exp3_v2/EXP3_V2_REV002_EXTERNAL_DRIVER_ATTEMPT_LOG_ARCHIVE.json",
    "phase_b/exp3_v2/EXP3_V2_REV002_EXTERNAL_DRIVER_FAILURE_ARCHIVE.json",
    "phase_b/exp3_v2/EXP3_V2_REV002_OFFICIAL_WRAPPER_ATTEMPT_LOG_ARCHIVE.json",
    "phase_b/exp3_v2/EXP3_V2_REV002_OFFICIAL_WRAPPER_FAILURE_ARCHIVE.json",
    "phase_b/exp3_v2/EXP3_V2_REV003_PREEXECUTION_INCIDENTS.json",
    "phase_b/exp3_v2/EXP3_V2_REV003_PREEXECUTION_INCIDENTS.md",
    "phase_b/exp3_v2/EXP3_V2_REV003_SENTINEL_ATTEMPT_LOG_ARCHIVE.json",
    "phase_b/exp3_v2/EXP3_V2_REV003_SENTINEL_FAILURE.json",
    "phase_b/exp3_v2/EXP3_V2_REV003_SENTINEL_FAILURE.md",
    "phase_b/exp3_v2/EXP3_V2_REV003_SENTINEL_FAILURE_ARCHIVE.json",
    "phase_b/exp3_v2/EXP3_V2_REV003_SENTINEL_MANIFEST_ARCHIVE.csv",
    "phase_b/exp3_v2/EXP3_V2_SENTINEL_PREFLIGHT_ABORT_001.json",
    "phase_b/exp3_v2/EXP3_V2_SENTINEL_PREFLIGHT_ABORT_001.md",
    "phase_b/exp3_v2/append_exp3v2_attempt_record.m",
    "phase_b/exp3_v2/assert_exp3v2_freeze_boundary.m",
    "phase_b/exp3_v2/assert_exp3v2_runtime_bundle.m",
    "phase_b/exp3_v2/configure_exp3v2_model.m",
    "phase_b/exp3_v2/configure_exp3v2_file_generation.m",
    "phase_b/exp3_v2/exp3v2_attempt_log.schema.json",
    "phase_b/exp3_v2/exp3v2_attempt_log.template.json",
    "phase_b/exp3_v2/exp3v2_case_plan.json",
    "phase_b/exp3_v2/exp3v2_manifest_template.csv",
    "phase_b/exp3_v2/exp3v2_sentinel_case.json",
    "phase_b/exp3_v2/exp3v2_shell_quote.m",
    "phase_b/exp3_v2/exp3v2_workspace_outputs_present.m",
    "phase_b/exp3_v2/extract_exp3v2_outputs.m",
    "phase_b/exp3_v2/format_exp3v2_csv_scalar.m",
    "phase_b/exp3_v2/generate_exp3v2_heldout.m",
    "phase_b/exp3_v2/generate_exp3v2_sentinel.m",
    "phase_b/exp3_v2/materialize_exp3v2_runtime.py",
    "phase_b/exp3_v2/restore_exp3v2_model_config.m",
    "phase_b/exp3_v2/restore_exp3v2_file_generation.m",
    "phase_b/exp3_v2/run_exp3v2_engine.m",
    "phase_b/exp3_v2/sentinel_integration_run.m",
    "phase_b/exp3_v2/test_exp3v2_attempt_policy.m",
    "phase_b/exp3_v2/test_exp3v2_file_generation_isolation.m",
    "phase_b/exp3_v2/test_exp3v2_manifest_contract.m",
    "phase_b/exp3_v2/test_exp3v2_model_config_management.m",
    "phase_b/exp3_v2/test_exp3v2_output_retrieval.m",
    "phase_b/exp3_v2/test_exp3v2_csv_serialization.m",
    "phase_b/exp3_v2/test_exp3v2_python_runtime_preflight.m",
    "phase_b/exp3_v2/test_exp3v2_runtime_provenance.m",
    "phase_b/exp3_v2/test_exp3v2_workspace_isolation.m",
    "phase_b/exp3_v2/verify_exp3v2_heldout.py",
    "phase_b/exp3_v2/validate_exp3v2_python_runtime.m",
    "phase_b/exp3_v2/write_exp3v2_sentinel_manifest.m",
    "phase_b/tests/test_exp3v2_pre_freeze.py",
    "phase_b/tests/test_exp3v2_runtime_materialization.py",
    "phase_b/PHASE_B_PROTOCOL_HASHES.json",
}
REQUIRED_HARNESS_PATHS.update(
    json.loads((ROOT / "phase_b/PHASE_B_PROTOCOL_HASHES.json").read_text())["artifacts"]
)
FINAL_REQUIRED_PATHS = REQUIRED_HARNESS_PATHS | {
    "phase_b/exp3_v2/EXP3_V2_HARNESS_FREEZE_MANIFEST_004.json",
    "phase_b/exp3_v2/EXP3_V2_SENTINEL_EVIDENCE.json",
    "phase_b/exp3_v2/EXP3_V2_SENTINEL_EVIDENCE.md",
}
FINAL_REVISION_002_REQUIRED_PATHS = FINAL_REQUIRED_PATHS | {
    "phase_b/exp3_v2/EXP3_V2_FREEZE_MANIFEST.json",
    "phase_b/exp3_v2/EXP3_V2_FINAL_BOUNDARY_REV002_PREFLIGHT_FAILURE.json",
    "phase_b/exp3_v2/EXP3_V2_FINAL_BOUNDARY_REV002_PREFLIGHT_FAILURE.md",
    "phase_b/exp3_v2/test_exp3v2_real_runtime_preflight.m",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def close(left: float, right: float) -> bool:
    return math.isclose(left, right, rel_tol=0.0, abs_tol=TIME_TOLERANCE)


def git(*arguments: str) -> str:
    return subprocess.run(
        ["git", *arguments],
        cwd=ROOT,
        check=True,
        capture_output=True,
        text=True,
    ).stdout.strip()


def git_bytes(*arguments: str) -> bytes:
    return subprocess.run(
        ["git", *arguments], cwd=ROOT, check=True, capture_output=True
    ).stdout


def canonical_cases() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    ordinal = 0
    for condition in EXPECTED_CONDITIONS:
        prefix = "N" if condition == "Normal" else condition
        for run_index in range(1, 7):
            ordinal += 1
            primary = 320000 + ordinal
            rows.append(
                {
                    "physical_case_id": f"EXP3V2-{prefix}-{run_index:03d}",
                    "condition": condition,
                    "run_index": run_index,
                    "rng_algorithm": "twister",
                    "primary_seed": primary,
                    "replacement_seed": primary + 1_000_000,
                    "max_total_attempts": 2,
                }
            )
    return rows


def validate_history() -> list[str]:
    errors: list[str] = []
    for commit, tag in EXPECTED_HISTORY:
        try:
            if git("rev-parse", f"{tag}^{{}}") != commit:
                errors.append(f"immutable tag target mismatch: {tag}")
        except Exception as exc:
            errors.append(f"immutable tag cannot be resolved: {tag}: {exc}")
    for (parent, _), (child, _) in zip(EXPECTED_HISTORY, EXPECTED_HISTORY[1:]):
        result = subprocess.run(
            ["git", "merge-base", "--is-ancestor", parent, child], cwd=ROOT
        )
        if result.returncode != 0:
            errors.append(f"immutable EXP3 ancestry mismatch: {parent} -> {child}")

    live = ROOT / "tep_exp3_heldout/exp3_attempt_log.json"
    archive = ROOT / "phase_b/exp3/EXP3_CLOSURE_attempt_log_archive.json"
    if not archive.is_file() or sha256_file(archive) != ATTEMPT_LOG_HASH:
        errors.append("EXP3 archived attempt-log hash mismatch")
    if live.exists():
        if not live.is_file() or sha256_file(live) != ATTEMPT_LOG_HASH:
            errors.append("EXP3 live attempt-log hash mismatch")
        elif archive.is_file() and live.read_bytes() != archive.read_bytes():
            errors.append("EXP3 closure attempt-log archive is not verbatim")

    try:
        manifest = json.loads(
            (ROOT / "phase_b/PHASE_B_PROTOCOL_HASHES.json").read_text()
        )
        artifacts = manifest["artifacts"]
        if len(artifacts) != 56:
            errors.append(f"Experiment 1 artifact count is {len(artifacts)}, not 56")
        for relative, expected in artifacts.items():
            path = ROOT / relative
            if not path.is_file() or sha256_file(path) != expected:
                errors.append(f"Experiment 1 frozen hash mismatch: {relative}")
    except Exception as exc:
        errors.append(f"Experiment 1 hash validation failed: {exc}")
    return errors


def validate_external_dependency_inventory(manifest: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    dependencies = manifest.get("external_runtime_dependencies")
    if not isinstance(dependencies, list):
        return ["external_runtime_dependencies must be an array"]
    required_fields = {"path", "size_bytes", "sha256", "role", "provenance"}
    observed: dict[str, tuple[Any, Any]] = {}
    for dependency in dependencies:
        if not isinstance(dependency, dict) or set(dependency) != required_fields:
            errors.append("external dependency entry schema mismatch")
            continue
        relative = dependency.get("path")
        if not isinstance(relative, str):
            errors.append("external dependency path must be a string")
            continue
        path = Path(relative)
        if path.is_absolute() or ".." in path.parts:
            errors.append(f"unsafe external dependency path: {relative}")
            continue
        if relative in observed:
            errors.append(f"duplicate external dependency path: {relative}")
        observed[relative] = (
            dependency.get("size_bytes"),
            dependency.get("sha256"),
        )
        if not isinstance(dependency.get("role"), str) or not dependency["role"]:
            errors.append(f"external dependency role missing: {relative}")
        if (
            not isinstance(dependency.get("provenance"), str)
            or not dependency["provenance"]
        ):
            errors.append(f"external dependency provenance missing: {relative}")
    if set(observed) != set(EXPECTED_EXTERNAL_DEPENDENCIES):
        errors.append("external dependency path set mismatch")
    for relative, expected in EXPECTED_EXTERNAL_DEPENDENCIES.items():
        if observed.get(relative) != expected:
            errors.append(f"external dependency size/hash mismatch: {relative}")
    return errors


def validate_runtime_directory(
    manifest: dict[str, Any], runtime_dir: Path
) -> list[str]:
    errors = validate_external_dependency_inventory(manifest)
    if errors:
        return errors
    if runtime_dir.is_symlink() or not runtime_dir.is_dir():
        return ["materialized runtime directory is missing or a symlink"]
    expected = manifest["external_runtime_dependencies"]
    expected_paths = {dependency["path"] for dependency in expected}
    for path in runtime_dir.rglob("*"):
        if path.is_symlink():
            errors.append(
                f"materialized runtime contains symlink: {path.relative_to(runtime_dir)}"
            )
    observed_paths = {
        path.relative_to(runtime_dir).as_posix()
        for path in runtime_dir.rglob("*")
        if path.is_file()
    }
    if observed_paths != expected_paths:
        errors.append("materialized runtime has missing or extra files")
    for dependency in expected:
        path = runtime_dir / dependency["path"]
        if path.is_symlink() or not path.is_file():
            errors.append(f"missing regular runtime dependency: {dependency['path']}")
            continue
        if path.stat().st_size != dependency["size_bytes"]:
            errors.append(f"runtime dependency size mismatch: {dependency['path']}")
        elif sha256_file(path) != dependency["sha256"]:
            errors.append(f"runtime dependency hash mismatch: {dependency['path']}")
    return errors


def validate_case_plan(plan: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if plan.get("status") not in {"PRE_FREEZE_DRAFT", "FROZEN_BEFORE_GENERATION"}:
        errors.append("case-plan status is invalid")
    if plan.get("experiment") != "Experiment 3 V2 — Prospective Fresh-Run Held-Out":
        errors.append("case-plan experiment identity mismatch")
    if plan.get("cases") != canonical_cases():
        errors.append("cases do not exactly match the canonical 30-case plan")
    cases = plan.get("cases", [])
    if Counter(row.get("condition") for row in cases) != Counter(
        {condition: 6 for condition in EXPECTED_CONDITIONS}
    ):
        errors.append("case-plan condition counts mismatch")
    primary = [row.get("primary_seed") for row in cases]
    replacement = [row.get("replacement_seed") for row in cases]
    if set(primary) & set(replacement):
        errors.append("primary/replacement seed collision")
    forbidden = set(range(310001, 310031)) | set(range(1310001, 1310031))
    forbidden |= {987654321, 987654322, 123456789, 320031}
    if (set(primary) | set(replacement)) & forbidden:
        errors.append("V2 run seeds collide with EXP3, sentinel, or bootstrap seeds")
    expected_rng = {
        "algorithm": "twister",
        "master_allocation_base": 320000,
        "primary_seed_formula": "master_allocation_base + canonical_ordinal",
        "primary_seed_range": [320001, 320030],
        "replacement_seed_formula": "primary_seed + 1000000",
        "allowed_attempts": [0, 1],
        "max_total_attempts": 2,
        "bootstrap_seed": 320031,
    }
    rng = plan.get("rng", {})
    for field, expected in expected_rng.items():
        if rng.get(field) != expected:
            errors.append(f"rng.{field} mismatch")
    if not isinstance(rng.get("fresh_allocation_rationale"), str):
        errors.append("fresh seed allocation rationale missing")
    if plan.get("runtime") != EXPECTED_RUNTIME:
        errors.append("runtime field set or five-field semantics mismatch")
    simulator = plan.get("simulator", {})
    for field, expected in EXPECTED_SIMULATOR.items():
        observed = simulator.get(field)
        if isinstance(expected, float):
            if not isinstance(observed, (int, float)) or not close(observed, expected):
                errors.append(f"simulator.{field} mismatch")
        elif observed != expected:
            errors.append(f"simulator.{field} mismatch")
    expected_statistics = {
        "independent_unit": "physical_run",
        "agent_case_observations_independent": False,
        "bootstrap": "paired_cluster_bootstrap_stratified_by_true_fault",
        "bootstrap_draws": 10000,
        "bootstrap_seed": 320031,
        "primary_analysis": "Experiment_3_V2_only",
        "primary_contrast": "B-A",
        "semantic_specificity_contrast": "B-E",
        "experiment_1_plus_experiment_3_v2": "secondary_descriptive_only",
    }
    if plan.get("statistics") != expected_statistics:
        errors.append("statistics block mismatch")
    return errors


def validate_sentinel_descriptor(
    descriptor: dict[str, Any], plan: dict[str, Any]
) -> list[str]:
    errors: list[str] = []
    expected = {
        "status": "SENTINEL_VALIDATION_ONLY",
        "physical_case_id": ACTIVE_SENTINEL_ID,
        "condition": "Normal",
        "attempt": 0,
        "seed": ACTIVE_SENTINEL_SEED,
        "rng_algorithm": "twister",
        "replacement_allowed": False,
    }
    for field, value in expected.items():
        if descriptor.get(field) != value:
            errors.append(f"sentinel descriptor {field} mismatch")
    ids = {row["physical_case_id"] for row in plan["cases"]}
    seeds = {plan["rng"]["bootstrap_seed"]}
    seeds |= {row["primary_seed"] for row in plan["cases"]}
    seeds |= {row["replacement_seed"] for row in plan["cases"]}
    if descriptor.get("physical_case_id") in ids or descriptor.get("seed") in seeds:
        errors.append("sentinel identity or seed collides with real allocation")
    consumed = descriptor.get("consumed_sentinels")
    if not isinstance(consumed, list) or len(consumed) != 1:
        errors.append("consumed sentinel history mismatch")
    else:
        entry = consumed[0]
        if (
            entry.get("physical_case_id") != "EXP3V2-SENTINEL-001"
            or entry.get("seed") != 987654321
            or entry.get("harness_revision") != "003"
            or entry.get("eligible_for_reuse") is not False
            or entry.get("outcome") != "END_TO_END_VERIFICATION_FAILED"
        ):
            errors.append("consumed sentinel entry mismatch")
    if (
        descriptor.get("physical_case_id") in CONSUMED_SENTINELS
        or descriptor.get("seed") in CONSUMED_SENTINELS.values()
    ):
        errors.append("active sentinel reuses a consumed identity or seed")
    return errors


def load_manifest(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as stream:
        reader = csv.DictReader(stream)
        if tuple(reader.fieldnames or ()) != MANIFEST_FIELDS:
            raise ValueError("manifest columns do not exactly match the template")
        return list(reader)


def validate_manifest_template(rows: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    expected = canonical_cases()
    if len(rows) != 30:
        return [f"manifest template has {len(rows)} rows, expected 30"]
    for row, case in zip(rows, expected):
        if row["physical_case_id"] != case["physical_case_id"]:
            errors.append("manifest template case order/ID mismatch")
        if row["fault/status"] != case["condition"]:
            errors.append(f"{case['physical_case_id']} condition mismatch")
        if row["seed"] != str(case["primary_seed"]):
            errors.append(f"{case['physical_case_id']} primary seed mismatch")
        for field in MANIFEST_FIELDS:
            if field not in {"physical_case_id", "fault/status", "seed"}:
                if row[field] != "TBD":
                    errors.append(f"{case['physical_case_id']} {field} must be TBD")
    return errors


def validate_freeze_manifest(
    path: Path,
    allowed_statuses: set[str],
    required_paths: set[str] | None = None,
    *,
    historical_boundary: bool = False,
) -> tuple[list[str], dict[str, Any]]:
    errors: list[str] = []
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        return [f"freeze manifest cannot be read: {exc}"], {}
    if payload.get("status") not in allowed_statuses:
        errors.append("freeze manifest status mismatch")
    if payload.get("hash_algorithm") != "SHA-256":
        errors.append("freeze manifest hash algorithm mismatch")
    if not isinstance(payload.get("supersedes"), dict):
        errors.append("freeze manifest supersedes contract missing")
    artifacts = payload.get("artifacts")
    if not isinstance(artifacts, list):
        return errors + ["freeze manifest artifacts must be an array"], payload
    if payload.get("artifact_count") != len(artifacts):
        errors.append("freeze manifest artifact count mismatch")
    paths = [row.get("path") for row in artifacts if isinstance(row, dict)]
    if len(paths) != len(set(paths)):
        errors.append("freeze manifest has duplicate artifact paths")
    if required_paths is not None and set(paths) != required_paths:
        errors.append("harness freeze artifact path set mismatch")
    if path.name == "EXP3_V2_HARNESS_FREEZE_MANIFEST.json":
        if payload.get("freeze_tag") != "exp3-v2-harness-frozen":
            errors.append("harness freeze tag mismatch")
        if payload.get("created_before_sentinel") is not True:
            errors.append("harness manifest is not pre-sentinel")
        if payload.get("sentinel_runs_at_freeze") != 0:
            errors.append("harness manifest sentinel count must be zero")
    if path.name == "EXP3_V2_HARNESS_FREEZE_MANIFEST_002.json":
        if payload.get("schema_version") != "2.0":
            errors.append("revision 002 schema version mismatch")
        if payload.get("manifest_revision") != "002":
            errors.append("harness manifest revision mismatch")
        if payload.get("freeze_tag") != "exp3-v2-harness-frozen-002":
            errors.append("revision 002 harness freeze tag mismatch")
        if payload.get("created_before_sentinel") is not True:
            errors.append("revision 002 manifest is not pre-sentinel")
        if payload.get("sentinel_runs_at_revision_preparation") != 0:
            errors.append("revision 002 sentinel count must be zero")
        if payload.get("status") == "PRE_FREEZE_DRAFT":
            if payload.get("tag_created") is not False:
                errors.append("draft revision 002 must record tag_created=false")
        errors.extend(validate_external_dependency_inventory(payload))
    if path.name == "EXP3_V2_HARNESS_FREEZE_MANIFEST_003.json":
        if payload.get("schema_version") != "3.0":
            errors.append("revision 003 schema version mismatch")
        if payload.get("manifest_revision") != "003":
            errors.append("revision 003 harness manifest revision mismatch")
        if payload.get("freeze_tag") != "exp3-v2-harness-frozen-003":
            errors.append("revision 003 harness freeze tag mismatch")
        if payload.get("created_before_sentinel") is not True:
            errors.append("revision 003 manifest is not pre-sentinel")
        if payload.get("sentinel_executions_at_revision_preparation") != 0:
            errors.append("revision 003 sentinel count must be zero")
        if payload.get("status") == "PRE_FREEZE_DRAFT":
            if payload.get("tag_created") is not False:
                errors.append("draft revision 003 must record tag_created=false")
        parent = payload.get("parent_revision_002")
        if not isinstance(parent, dict) or parent != {
            "commit": "261e54b10fe2c0a8897627ff7626c1a2d05672f8",
            "tag": "exp3-v2-harness-frozen-002",
            "manifest_sha256": (
                "c552a6f474491243f549f9588eec52d61fe65922ef8734ff843ef75745710019"
            ),
        }:
            errors.append("revision 003 parent provenance mismatch")
        if payload.get("case_plan_sha256") != (
            "3d102383b9eb8d5de14bffef862c2b5715d8bbcf05359decb5fdf31efe31a014"
        ):
            errors.append("revision 003 case-plan hash mismatch")
        if payload.get("unavailable_incident_evidence") != []:
            errors.append("revision 003 unavailable-evidence disclosure mismatch")
        errors.extend(validate_external_dependency_inventory(payload))
    if path.name == "EXP3_V2_HARNESS_FREEZE_MANIFEST_004.json":
        if payload.get("schema_version") != "4.0":
            errors.append("revision 004 schema version mismatch")
        if payload.get("manifest_revision") != "004":
            errors.append("revision 004 harness manifest revision mismatch")
        if payload.get("freeze_tag") != "exp3-v2-harness-frozen-004":
            errors.append("revision 004 harness freeze tag mismatch")
        if payload.get("created_before_sentinel") is not True:
            errors.append("revision 004 manifest is not pre-sentinel")
        if (
            payload.get("status") == "PRE_FREEZE_DRAFT"
            and payload.get("tag_created") is not False
        ):
            errors.append("draft revision 004 must record tag_created=false")
        if (
            payload.get("status") == "HARNESS_FROZEN_FOR_SENTINEL"
            and payload.get("tag_created") is not True
        ):
            errors.append("frozen revision 004 must record tag_created=true")
        parent = payload.get("parent_revision_003")
        if not isinstance(parent, dict) or parent != {
            "commit": "bce8f0e2f24db7033b7ddbecc38e1bfaa74c85a6",
            "tag": "exp3-v2-harness-frozen-003",
            "manifest_sha256": (
                "e9db49a5a71a4ffbb83213f81224e54569d85c841a31ca492c29a0fb32a62e03"
            ),
        }:
            errors.append("revision 004 parent provenance mismatch")
        if payload.get("case_plan_sha256") != (
            "3d102383b9eb8d5de14bffef862c2b5715d8bbcf05359decb5fdf31efe31a014"
        ):
            errors.append("revision 004 case-plan hash mismatch")
        if payload.get("active_sentinel") != {
            "physical_case_id": ACTIVE_SENTINEL_ID,
            "seed": ACTIVE_SENTINEL_SEED,
            "seed_consumed": False,
        }:
            errors.append("revision 004 active sentinel mismatch")
        if payload.get("consumed_sentinels") != [
            {
                "physical_case_id": "EXP3V2-SENTINEL-001",
                "seed": 987654321,
                "harness_revision": "003",
                "rng_calls": 1,
                "sim_calls": 1,
                "eligible_for_reuse": False,
            }
        ]:
            errors.append("revision 004 consumed sentinel history mismatch")
        if payload.get("sentinel_executions_at_revision_preparation") != 0:
            errors.append("revision 004 new sentinel count must be zero")
        if payload.get("rng_seed_calls_at_revision_preparation") != 0:
            errors.append("revision 004 preparation RNG count must be zero")
        if payload.get("sim_calls_at_revision_preparation") != 0:
            errors.append("revision 004 preparation sim count must be zero")
        if payload.get("workbooks_created_at_revision_preparation") != 0:
            errors.append("revision 004 preparation workbook count must be zero")
        errors.extend(validate_external_dependency_inventory(payload))
    if path.name == "EXP3_V2_FREEZE_MANIFEST.json":
        if payload.get("freeze_tag") != "exp3-v2-heldout-frozen":
            errors.append("final freeze tag mismatch")
        if (
            payload.get("status") == FINAL_CANDIDATE_STATUS
            and payload.get("tag_created") is not False
        ):
            errors.append("final review candidate must record tag_created=false")
        if (
            payload.get("status") == "FROZEN_BEFORE_GENERATION"
            and payload.get("tag_created") is not True
        ):
            errors.append("final frozen manifest must record tag_created=true")
        if payload.get("status") == "FROZEN_BEFORE_GENERATION":
            if payload.get("freeze_human_approval") != (
                "APPROVO IL FINAL FREEZE EXP3_V2"
            ):
                errors.append("final freeze human approval mismatch")
            if payload.get("freeze_human_approval_date") != "2026-09-03":
                errors.append("final freeze human approval date mismatch")
        if payload.get("created_before_real_generation") is not True:
            errors.append("final manifest is not pre-generation")
        if payload.get("sentinel_validation_passed") is not True:
            errors.append("final manifest does not bind sentinel PASS")
        if payload.get("v2_workbooks_at_freeze") != 0:
            errors.append("final manifest workbook count must be zero")
        if payload.get("real_attempt_log_present") is not False:
            errors.append("final manifest real attempt-log state mismatch")
        if payload.get("scientific_seeds_consumed") != 0:
            errors.append("final manifest scientific seed count must be zero")
        if not artifacts:
            errors.append("final freeze manifest artifacts are empty")
        if payload.get("harness_artifact_count") != len(REQUIRED_HARNESS_PATHS):
            errors.append("final manifest harness artifact count mismatch")
        if payload.get("harness_boundary") != {
            "revision": "004",
            "tag": "exp3-v2-harness-frozen-004",
            "commit": "258f629f07aad84b6186381fa6a1dab52401bd2f",
            "manifest_sha256": (
                "dacc810bd29203d3d701e3613a9ce8c72dc6423aa475f5c4e8c8b4989b40e139"
            ),
        }:
            errors.append("final manifest harness boundary mismatch")
        if payload.get("sentinel_evidence") != {
            "physical_case_id": ACTIVE_SENTINEL_ID,
            "seed": ACTIVE_SENTINEL_SEED,
            "json_path": "phase_b/exp3_v2/EXP3_V2_SENTINEL_EVIDENCE.json",
            "json_sha256": SENTINEL_EVIDENCE_HASHES["EXP3_V2_SENTINEL_EVIDENCE.json"],
            "markdown_path": "phase_b/exp3_v2/EXP3_V2_SENTINEL_EVIDENCE.md",
            "markdown_sha256": SENTINEL_EVIDENCE_HASHES["EXP3_V2_SENTINEL_EVIDENCE.md"],
            "workbook_sha256": (
                "337f3e709554dfa95a52fd0bab8bedbacb71a400378bc709a22538380d94fbd6"
            ),
            "workbook_size_bytes": 1704419,
            "rows": 3001,
            "cols": 54,
            "verifier": "PASS",
            "round_trip": "PASS",
            "restoration": "PASS",
            "isolation": "PASS",
            "real_path_non_interference": "PASS",
            "throwaway_cleanup": "PASS",
        }:
            errors.append("final manifest sentinel evidence binding mismatch")
        if payload.get("case_plan") != {
            "path": "phase_b/exp3_v2/exp3v2_case_plan.json",
            "sha256": (
                "3d102383b9eb8d5de14bffef862c2b5715d8bbcf05359decb5fdf31efe31a014"
                if payload.get("status") == FINAL_CANDIDATE_STATUS
                else "84d5af21847033fe4a5924f42fca0fb772201116a9759d6d85f8356929f2b21e"
            ),
            "status": (
                "PRE_FREEZE_DRAFT"
                if payload.get("status") == FINAL_CANDIDATE_STATUS
                else "FROZEN_BEFORE_GENERATION"
            ),
            "bytes_unchanged_from_harness": (
                payload.get("status") == FINAL_CANDIDATE_STATUS
            ),
        }:
            errors.append("final manifest case-plan binding mismatch")
        pending = payload.get("status") == FINAL_CANDIDATE_STATUS
        if payload.get("finalization_pending") != {
            "human_review": pending,
            "case_plan_status_transition": pending,
            "commit": pending,
            "annotated_tag": pending,
        }:
            errors.append("final manifest pending-finalization contract mismatch")
    if path.name == "EXP3_V2_FREEZE_MANIFEST_002.json":
        if payload.get("schema_version") != "3.0":
            errors.append("final boundary revision 002 schema mismatch")
        if payload.get("manifest_revision") != "002":
            errors.append("final boundary revision mismatch")
        if payload.get("status") not in {
            FINAL_REVISION_DRAFT_STATUS,
            "FROZEN_BEFORE_GENERATION",
        }:
            errors.append("final boundary revision 002 status mismatch")
        if payload.get("freeze_tag") != "exp3-v2-heldout-frozen-002":
            errors.append("final boundary revision 002 tag mismatch")
        if (
            payload.get("status") == FINAL_REVISION_DRAFT_STATUS
            and payload.get("tag_created") is not False
        ):
            errors.append("draft final boundary must record tag_created=false")
        if (
            payload.get("status") == "FROZEN_BEFORE_GENERATION"
            and payload.get("tag_created") is not True
        ):
            errors.append("frozen final boundary must record tag_created=true")
        if payload.get("status") == "FROZEN_BEFORE_GENERATION":
            if payload.get("freeze_human_approval") != (
                "APPROVO IL FREEZE EXP3_V2 FINAL BOUNDARY REVISION 002"
            ):
                errors.append("revision 002 freeze approval mismatch")
            if payload.get("freeze_human_approval_date") != "2026-09-03":
                errors.append("revision 002 freeze approval date mismatch")
        if payload.get("created_before_real_generation") is not True:
            errors.append("revision 002 is not before real generation")
        if payload.get("scientific_seeds_consumed") != 0:
            errors.append("revision 002 scientific seed count must be zero")
        if payload.get("real_workbooks_created") != 0:
            errors.append("revision 002 real workbook count must be zero")
        if payload.get("v2_workbooks_at_freeze") != 0:
            errors.append("revision 002 frozen workbook count must be zero")
        if payload.get("sentinel_validation_passed") is not True:
            errors.append("revision 002 does not preserve sentinel PASS")
        if payload.get("attempt_log_present") is not False:
            errors.append("revision 002 attempt-log state mismatch")
        if payload.get("original_final_boundary") != {
            "tag": "exp3-v2-heldout-frozen",
            "commit": "a55537dfc85db7e70f32ada21afffcb4e8824b96",
            "manifest_path": "phase_b/exp3_v2/EXP3_V2_FREEZE_MANIFEST.json",
            "manifest_sha256": ORIGINAL_FINAL_MANIFEST_HASH,
            "status": "FROZEN_BEFORE_GENERATION",
        }:
            errors.append("revision 002 original final-boundary provenance mismatch")
        original = json.loads((SCRIPT_DIR / "EXP3_V2_FREEZE_MANIFEST.json").read_text())
        if sha256_file(SCRIPT_DIR / "EXP3_V2_FREEZE_MANIFEST.json") != (
            ORIGINAL_FINAL_MANIFEST_HASH
        ):
            errors.append("original final manifest was modified")
        if payload.get("harness_boundary") != original.get("harness_boundary"):
            errors.append("revision 002 harness boundary mismatch")
        if payload.get("sentinel_evidence") != original.get("sentinel_evidence"):
            errors.append("revision 002 sentinel evidence mismatch")
        if payload.get("case_plan") != {
            "path": "phase_b/exp3_v2/exp3v2_case_plan.json",
            "sha256": (
                "84d5af21847033fe4a5924f42fca0fb772201116a9759d6d85f8356929f2b21e"
            ),
            "status": "FROZEN_BEFORE_GENERATION",
            "scientific_allocations_unchanged": True,
        }:
            errors.append("revision 002 case-plan binding mismatch")
        if payload.get("preflight_failure") != {
            "physical_case_id": "EXP3V2-N-001",
            "attempt": 0,
            "primary_seed": 320001,
            "failure_identifier": "EXP3V2:RuntimeDependencyManifest",
            "rng_calls": 0,
            "sim_calls": 0,
            "attempt_log_present": False,
            "workbooks_created": 0,
            "only_empty_output_directories_created": True,
            "seed_consumed": False,
            "attempt_0_remains_eligible": True,
            "json_path": (
                "phase_b/exp3_v2/"
                "EXP3_V2_FINAL_BOUNDARY_REV002_PREFLIGHT_FAILURE.json"
            ),
            "json_sha256": (
                "7e46bcaa77748bcbea08c0246eaee3d235609dc580a2dce69b74be01047c7233"
            ),
            "markdown_path": (
                "phase_b/exp3_v2/" "EXP3_V2_FINAL_BOUNDARY_REV002_PREFLIGHT_FAILURE.md"
            ),
            "markdown_sha256": (
                "5f3c00228419df90e22890d32ad647ced626dd633438148a77c3c6af03d21faf"
            ),
        }:
            errors.append("revision 002 preflight-failure binding mismatch")
        if payload.get("shared_engine") != {
            "path": "phase_b/exp3_v2/run_exp3v2_engine.m",
            "sha256": (
                "4e746a8b6504953d2bb0d4eb9982cdef1ee3c02d0d0f1cb374e5a9086e45a9f1"
            ),
            "modified": False,
        }:
            errors.append("revision 002 shared-engine immutability mismatch")
        pending = payload.get("status") == FINAL_REVISION_DRAFT_STATUS
        if payload.get("finalization_pending") != {
            "human_review": pending,
            "commit": pending,
            "annotated_tag": pending,
            "tag_push": pending,
        }:
            errors.append("revision 002 pending-finalization contract mismatch")
        errors.extend(validate_external_dependency_inventory(payload))
        if payload.get("status") == FINAL_REVISION_DRAFT_STATUS:
            try:
                git("rev-parse", "exp3-v2-heldout-frozen-002^{}")
                errors.append("prospective final boundary revision 002 tag exists")
            except Exception:
                pass
    manifest_relative = path.resolve().relative_to(ROOT).as_posix()
    if manifest_relative in paths:
        errors.append("freeze manifest illegally contains its own hash")
    for row in artifacts:
        if not isinstance(row, dict) or set(row) != {"path", "sha256"}:
            errors.append("freeze manifest artifact entry schema mismatch")
            continue
        relative = Path(row["path"])
        digest = row["sha256"]
        if relative.is_absolute() or ".." in relative.parts:
            errors.append(f"unsafe freeze artifact path: {relative}")
        elif not isinstance(digest, str) or not HASH_RE.fullmatch(digest):
            errors.append(f"invalid freeze artifact hash: {relative}")
        elif not (ROOT / relative).is_file():
            errors.append(f"missing freeze artifact: {relative}")
        elif not historical_boundary and sha256_file(ROOT / relative) != digest:
            errors.append(f"freeze artifact hash mismatch: {relative}")
        if relative.as_posix() == "tep_exp3_heldout/exp3_attempt_log.json" or (
            relative.parts[:2] == ("tep_parent_a0413e16", "simulator")
        ):
            errors.append(f"external or live ignored file in Git boundary: {relative}")
    status = payload.get("status")
    if status in {"HARNESS_FROZEN_FOR_SENTINEL", "FROZEN_BEFORE_GENERATION"}:
        tag = payload.get("freeze_tag")
        try:
            target = git("rev-parse", f"{tag}^{{}}")
            historical_targets = {
                "EXP3_V2_HARNESS_FREEZE_MANIFEST_004.json": (
                    "258f629f07aad84b6186381fa6a1dab52401bd2f"
                ),
                "EXP3_V2_FREEZE_MANIFEST.json": (
                    "a55537dfc85db7e70f32ada21afffcb4e8824b96"
                ),
            }
            expected_target = (
                historical_targets.get(path.name, git("rev-parse", "HEAD"))
                if historical_boundary
                else git("rev-parse", "HEAD")
            )
            if target != expected_target:
                errors.append("freeze tag target does not equal expected commit")
        except Exception as exc:
            errors.append(f"freeze tag cannot be verified: {exc}")
        artifact_hashes = {
            row["path"]: row["sha256"]
            for row in artifacts
            if isinstance(row, dict) and set(row) == {"path", "sha256"}
        }
        for relative in paths:
            try:
                tagged_bytes = git_bytes("show", f"{tag}:{relative}")
                comparison_hash = (
                    artifact_hashes.get(relative)
                    if historical_boundary
                    else sha256_file(ROOT / relative)
                )
                if hashlib.sha256(tagged_bytes).hexdigest() != comparison_hash:
                    errors.append(f"Git tree/worktree artifact mismatch: {relative}")
            except Exception as exc:
                errors.append(
                    f"artifact is not materialized by Git tree: {relative}: {exc}"
                )
    return errors, payload


def validate_generator_contract(manifest: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    engine = (SCRIPT_DIR / "run_exp3v2_engine.m").read_text()
    extractor = (SCRIPT_DIR / "extract_exp3v2_outputs.m").read_text()
    wrappers = "\n".join(
        (SCRIPT_DIR / name).read_text()
        for name in ("generate_exp3v2_heldout.m", "generate_exp3v2_sentinel.m")
    )
    if (
        len(
            re.findall(
                r"rng\(seed, 'twister'\);\s*\n\s*simResult = sim\(modelName\);", engine
            )
        )
        != 1
    ):
        errors.append("shared engine does not preserve exact rng-to-sim adjacency")
    if len(re.findall(r"\bsim\s*\(", engine)) != 1:
        errors.append("shared engine must contain exactly one sim call")
    forbidden = (
        "evalin('base', 'tout')",
        "evalin('base', 'simout')",
        "evalin('base', 'xmv')",
    )
    if any(token in engine or token in extractor for token in forbidden):
        errors.append("base-workspace output fallback is present")
    for token in (
        "isa(simResult, 'Simulink.SimulationOutput')",
        "who(simResult)",
        "simResult.get('tout')",
        "simResult.get('simout')",
        "simResult.get('xmv')",
        "[expectedRows 1]",
        "[expectedRows 41]",
        "[expectedRows 12]",
    ):
        if token not in extractor:
            errors.append(f"typed output retrieval contract missing: {token}")
    for token in (
        "configure_exp3v2_model(modelName)",
        "restore_exp3v2_model_config(configState)",
        "ReturnWorkspaceOutputs",
        "clear tout simout xmv",
        "exp3v2_workspace_outputs_present()",
    ):
        if token not in engine:
            errors.append(f"shared-engine safety contract missing: {token}")
    if "save_system" in engine or "save_system" in wrappers:
        errors.append("generator must never save the model")
    fields = set(
        re.findall(r"(?:freezeManifest|harnessManifest)\.([A-Za-z0-9_]+)", wrappers)
    )
    revision_path = SCRIPT_DIR / "EXP3_V2_FREEZE_MANIFEST_002.json"
    final_path = (
        revision_path
        if revision_path.is_file()
        else SCRIPT_DIR / "EXP3_V2_FREEZE_MANIFEST.json"
    )
    final_manifest = json.loads(final_path.read_text())
    missing = fields - (set(manifest) | set(final_manifest))
    if missing:
        errors.append(f"generator/manifest fields missing: {sorted(missing)}")
    return errors


def load_attempt_log(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("attempts"), list):
        raise ValueError("attempt log must contain an attempts array")
    return payload


def validate_attempt_log(
    payload: dict[str, Any], plan: dict[str, Any], schema: dict[str, Any]
) -> list[str]:
    errors: list[str] = []
    validator = Draft202012Validator(schema, format_checker=FormatChecker())
    for error in validator.iter_errors(payload):
        errors.append(f"attempt-log schema: {error.message}")
    attempts = payload.get("attempts", [])
    by_case = {row["physical_case_id"]: row for row in plan["cases"]}
    seen: set[tuple[str, int]] = set()
    prior: dict[str, dict[int, dict[str, Any]]] = {}
    for position, row in enumerate(attempts):
        label = f"attempts[{position}]"
        if not isinstance(row, dict):
            errors.append(f"{label} must be an object")
            continue
        if set(row) != ATTEMPT_FIELDS:
            errors.append(f"{label} technical-provenance field set mismatch")
            continue
        case_id = row["physical_case_id"]
        attempt = row["attempt"]
        if case_id not in by_case or type(attempt) is not int or attempt not in {0, 1}:
            errors.append(f"{label} unknown case or invalid attempt")
            continue
        key = (case_id, attempt)
        if key in seen:
            errors.append(f"duplicate attempt: {key}")
        seen.add(key)
        case = by_case[case_id]
        expected_seed = (
            case["primary_seed"] if attempt == 0 else case["replacement_seed"]
        )
        if row["seed"] != expected_seed:
            errors.append(f"{label} seed mismatch")
        if row["fault_status"] != case["condition"]:
            errors.append(f"{label} fault/status mismatch")
        for field, value in EXPECTED_RUNTIME.items():
            if row[field] != value:
                errors.append(f"{label} {field} mismatch")
        for field, value in EXPECTED_ATTEMPT_SIMULATOR.items():
            if row[field] != value:
                errors.append(f"{label} {field} mismatch")
        if row["case_plan_hash"] != sha256_file(SCRIPT_DIR / "exp3v2_case_plan.json"):
            errors.append(f"{label} case_plan_hash mismatch")
        if row["generation_script_hash"] != sha256_file(
            SCRIPT_DIR / "run_exp3v2_engine.m"
        ):
            errors.append(f"{label} generation_script_hash mismatch")
        output_path = Path(str(row["output_path"]))
        expected_name = f"{case_id}__attempt-{attempt}.xlsx"
        if output_path.name != expected_name or not path_within(output_path, REAL_ROOT):
            errors.append(f"{label} output path mismatch")
        indexed = prior.setdefault(case_id, {})
        if attempt == 1:
            primary = indexed.get(0)
            if primary is None:
                errors.append(f"{label} replacement has no earlier attempt 0")
            elif primary["structural_valid"]:
                errors.append(f"{label} replacement follows a valid attempt 0")
            elif not primary["technical_failure_reason"]:
                errors.append(f"{label} attempt 0 lacks technical failure reason")
        indexed[attempt] = row
        errors.extend(validate_structural_record(row, label))
    return errors


def validate_structural_record(row: dict[str, Any], label: str) -> list[str]:
    errors: list[str] = []
    if row.get("structural_valid") is True:
        expected = {
            "rows": 3001,
            "cols": 54,
            "finite_check": True,
            "technical_failure_reason": "",
        }
        for field, value in expected.items():
            if row.get(field) != value:
                errors.append(f"{label} {field} mismatch")
        if not close(float(row.get("time_start")), 0.0):
            errors.append(f"{label} time_start mismatch")
        if not close(float(row.get("time_end")), 50.0):
            errors.append(f"{label} time_end mismatch")
        if not close(float(row.get("sampling_interval")), 1 / 60):
            errors.append(f"{label} sampling mismatch")
        if row.get("output_size_bytes", 0) <= 0 or not HASH_RE.fullmatch(
            str(row.get("output_sha256", ""))
        ):
            errors.append(f"{label} materialized output provenance missing")
    elif not row.get("technical_failure_reason"):
        errors.append(f"{label} failed attempt lacks technical failure reason")
    return errors


def inspect_workbook(path: Path) -> dict[str, Any]:
    try:
        with ZipFile(path) as archive:
            if archive.testzip() is not None:
                raise ValueError("damaged XLSX ZIP container")
    except (BadZipFile, OSError) as exc:
        raise ValueError("invalid XLSX ZIP container") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if workbook.sheetnames != ["Sheet1"]:
            raise ValueError("worksheet mismatch")
        rows = workbook["Sheet1"].iter_rows(values_only=True)
        if list(next(rows)) != EXPECTED_HEADER:
            raise ValueError("header mismatch")
        times: list[float] = []
        finite = True
        count = 0
        for count, row in enumerate(rows, start=1):
            if len(row) != 54:
                raise ValueError("column count mismatch")
            for column, value in enumerate(row):
                if isinstance(value, bool) or not isinstance(value, (int, float)):
                    finite = False
                    continue
                finite = finite and math.isfinite(float(value))
                if column == 0:
                    times.append(float(value))
        intervals = [right - left for left, right in zip(times, times[1:])]
        return {
            "rows": count,
            "cols": workbook["Sheet1"].max_column,
            "time_start": times[0],
            "time_end": times[-1],
            "sampling": intervals[0],
            "finite_check": finite,
            "structural_valid": count == 3001
            and workbook["Sheet1"].max_column == 54
            and close(times[0], 0.0)
            and close(times[-1], 50.0)
            and all(value > 0 for value in intervals)
            and all(close(value, 1 / 60) for value in intervals)
            and finite,
        }
    finally:
        workbook.close()


def path_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def paths_overlap(left: Path, right: Path) -> bool:
    return path_within(left, right) or path_within(right, left)


def validate_materialized(
    rows: list[dict[str, str]], attempts: dict[str, Any], data_dir: Path
) -> list[str]:
    errors: list[str] = []
    indexed = {
        (row["physical_case_id"], row["attempt"]): row
        for row in attempts["attempts"]
        if isinstance(row, dict)
    }
    expected_files: set[str] = set()
    for row in rows:
        label = f"manifest[{row.get('physical_case_id')}]"
        try:
            attempt = int(row["attempt"])
            seed = int(row["seed"])
            size = int(row["size_bytes"])
            expected_rows = int(row["rows"])
            expected_cols = int(row["cols"])
            expected_start = float(row["time_start"])
            expected_end = float(row["time_end"])
            expected_sampling = float(row["sampling"])
        except (KeyError, TypeError, ValueError) as exc:
            errors.append(f"{label} invalid typed value: {exc}")
            continue
        filename = row["filename"]
        expected_files.add(filename)
        record = indexed.get((row["physical_case_id"], attempt))
        if record is None or not record.get("structural_valid"):
            errors.append(f"{label} lacks matching valid attempt")
            continue
        if record["seed"] != seed:
            errors.append(f"{label} seed disagrees with attempt log")
        expected_name = f"{row['physical_case_id']}__attempt-{attempt}.xlsx"
        if (
            filename != expected_name
            or Path(str(record["output_path"])).name != filename
        ):
            errors.append(f"{label} filename disagrees with identity/attempt log")
        comparisons = {
            "output_size_bytes": size,
            "output_sha256": row["SHA256"],
            "rows": expected_rows,
            "cols": expected_cols,
        }
        for field, value in comparisons.items():
            if record[field] != value:
                errors.append(f"{label} {field} disagrees with attempt log")
        path = data_dir / filename
        if Path(filename).name != filename or not path.is_file():
            errors.append(f"{label} workbook missing or unsafe filename")
            continue
        if path.stat().st_size != size or sha256_file(path) != row["SHA256"]:
            errors.append(f"{label} workbook size/hash mismatch")
        try:
            observed = inspect_workbook(path)
        except Exception as exc:
            errors.append(f"{label} workbook error: {exc}")
            continue
        expected_values = {
            "rows": expected_rows,
            "cols": expected_cols,
            "time_start": expected_start,
            "time_end": expected_end,
            "sampling": expected_sampling,
            "finite_check": row["finite_check"].lower() == "true",
            "structural_valid": row["structural_valid"].lower() == "true",
        }
        for field, value in expected_values.items():
            observed_value = observed[field]
            if isinstance(value, float):
                if not close(float(observed_value), value):
                    errors.append(f"{label} observed {field} mismatch")
            elif observed_value != value:
                errors.append(f"{label} observed {field} mismatch")
    observed_files = {path.name for path in data_dir.glob("*.xlsx")}
    logged_files = {
        Path(str(row["output_path"])).name
        for row in attempts["attempts"]
        if isinstance(row, dict) and int(row.get("output_size_bytes", 0)) > 0
    }
    if observed_files != expected_files or observed_files != logged_files:
        errors.append("data directory/manifest workbook set mismatch")
    return errors


def validate_sentinel_evidence() -> list[str]:
    errors: list[str] = []
    for name, expected_hash in SENTINEL_EVIDENCE_HASHES.items():
        path = SCRIPT_DIR / name
        if not path.is_file():
            errors.append(f"missing sentinel evidence: {name}")
        elif sha256_file(path) != expected_hash:
            errors.append(f"sentinel evidence hash mismatch: {name}")
    json_path = SCRIPT_DIR / "EXP3_V2_SENTINEL_EVIDENCE.json"
    if not json_path.is_file():
        return errors
    try:
        evidence = json.loads(json_path.read_text())
    except Exception as exc:
        return errors + [f"sentinel evidence cannot be read: {exc}"]
    expected = {
        "status": "PASS",
        "harness_revision": "004",
        "harness_tag": "exp3-v2-harness-frozen-004",
        "harness_commit": "258f629f07aad84b6186381fa6a1dab52401bd2f",
        "harness_manifest_sha256": (
            "dacc810bd29203d3d701e3613a9ce8c72dc6423aa475f5c4e8c8b4989b40e139"
        ),
        "physical_case_id": ACTIVE_SENTINEL_ID,
        "seed": ACTIVE_SENTINEL_SEED,
        "rng_algorithm": "twister",
        "workbook_sha256": (
            "337f3e709554dfa95a52fd0bab8bedbacb71a400378bc709a22538380d94fbd6"
        ),
        "workbook_size_bytes": 1704419,
        "rows": 3001,
        "cols": 54,
        "time_axis_check": "PASS",
        "finiteness_check": "PASS",
        "round_trip_check": "PASS",
        "verifier_output": "PASS: Experiment 3 V2 sentinel verification succeeded.",
        "runtime_materialization": "PASS",
        "throwaway_cleanup": "PASS",
    }
    for field, value in expected.items():
        if evidence.get(field) != value:
            errors.append(f"sentinel evidence {field} mismatch")
    for field in (
        "stopfcn_restored",
        "return_workspace_outputs_restored",
        "dirty_state_restored",
        "model_sha256_restored",
        "injected_error_restoration_test",
        "source_simulator_unchanged",
        "file_generation_isolated",
        "file_generation_configuration_restored",
        "real_path_non_interference",
    ):
        if evidence.get(field) is not True:
            errors.append(f"sentinel evidence {field} is not true")
    if evidence.get("python_runtime") != {
        "executable_path": "/opt/anaconda3/bin/python3.13",
        "python_version": "3.13.9",
        "jsonschema_version": "4.25.0",
        "openpyxl_version": "3.1.5",
    }:
        errors.append("sentinel evidence Python runtime mismatch")
    if evidence.get("runtime") != EXPECTED_RUNTIME:
        errors.append("sentinel evidence MATLAB/Simulink runtime mismatch")
    harness = json.loads(
        (SCRIPT_DIR / "EXP3_V2_HARNESS_FREEZE_MANIFEST_004.json").read_text()
    )
    if evidence.get("artifact_hashes") != harness.get("artifacts"):
        errors.append("sentinel evidence harness artifact hashes mismatch")
    if evidence.get("external_runtime_dependencies") != harness.get(
        "external_runtime_dependencies"
    ):
        errors.append("sentinel evidence external runtime inventory mismatch")
    return errors


def prefreeze_checks(
    case_plan_path: Path,
    harness_manifest_path: Path,
    final_manifest_path: Path,
    runtime_dir: Path | None = None,
) -> list[str]:
    errors: list[str] = []
    plan = json.loads(case_plan_path.read_text())
    final = json.loads(final_manifest_path.read_text())
    is_revision_002 = final_manifest_path.name == "EXP3_V2_FREEZE_MANIFEST_002.json"
    historical_harness = (
        final.get("status")
        in {
            FINAL_CANDIDATE_STATUS,
            "FROZEN_BEFORE_GENERATION",
        }
        or is_revision_002
    )
    errors.extend(validate_case_plan(plan))
    descriptor = json.loads((SCRIPT_DIR / "exp3v2_sentinel_case.json").read_text())
    errors.extend(validate_sentinel_descriptor(descriptor, plan))
    rows = load_manifest(SCRIPT_DIR / "exp3v2_manifest_template.csv")
    errors.extend(validate_manifest_template(rows))
    schema = json.loads((SCRIPT_DIR / "exp3v2_attempt_log.schema.json").read_text())
    Draft202012Validator.check_schema(schema)
    template = load_attempt_log(SCRIPT_DIR / "exp3v2_attempt_log.template.json")
    errors.extend(
        error.message for error in Draft202012Validator(schema).iter_errors(template)
    )
    if template["attempts"] != []:
        errors.append("attempt-log template must be empty")
    manifest_errors, harness = validate_freeze_manifest(
        harness_manifest_path,
        {"PRE_FREEZE_DRAFT", "HARNESS_FROZEN_FOR_SENTINEL"},
        REQUIRED_HARNESS_PATHS,
        historical_boundary=historical_harness,
    )
    errors.extend(manifest_errors)
    errors.extend(validate_generator_contract(harness))
    if runtime_dir is not None:
        runtime_manifest = final if is_revision_002 else harness
        errors.extend(validate_runtime_directory(runtime_manifest, runtime_dir))
    try:
        if final.get("status") not in {
            "PENDING_SENTINEL_VALIDATION",
            FINAL_CANDIDATE_STATUS,
            "FROZEN_BEFORE_GENERATION",
            FINAL_REVISION_DRAFT_STATUS,
        }:
            errors.append("final freeze manifest status mismatch")
        if not is_revision_002 and final.get("status") in {
            FINAL_CANDIDATE_STATUS,
            "FROZEN_BEFORE_GENERATION",
        }:
            final_errors, _ = validate_freeze_manifest(
                final_manifest_path, {final.get("status")}, FINAL_REQUIRED_PATHS
            )
            errors.extend(final_errors)
            errors.extend(validate_sentinel_evidence())
            frozen_harness_errors, frozen_harness = validate_freeze_manifest(
                harness_manifest_path,
                {"HARNESS_FROZEN_FOR_SENTINEL"},
                REQUIRED_HARNESS_PATHS,
                historical_boundary=True,
            )
            errors.extend(frozen_harness_errors)
            harness_hashes = {
                row["path"]: row["sha256"] for row in frozen_harness["artifacts"]
            }
            allowed = {
                row["path"]: row
                for row in final.get("allowed_finalization_changes", [])
            }
            observed_changes: set[str] = set()
            for relative, original_hash in harness_hashes.items():
                observed_hash = sha256_file(ROOT / relative)
                if observed_hash != original_hash:
                    observed_changes.add(relative)
                    change = allowed.get(relative)
                    if (
                        not isinstance(change, dict)
                        or change.get("before_sha256") != original_hash
                        or change.get("after_sha256") != observed_hash
                    ):
                        errors.append(
                            f"unreviewed harness artifact change after sentinel: {relative}"
                        )
            if set(allowed) != observed_changes:
                errors.append("allowed finalization change set mismatch")
            expected_plan_status = (
                "PRE_FREEZE_DRAFT"
                if final.get("status") == FINAL_CANDIDATE_STATUS
                else "FROZEN_BEFORE_GENERATION"
            )
            if plan.get("status") != expected_plan_status:
                errors.append("case-plan status does not match final-freeze phase")
        elif is_revision_002:
            final_errors, _ = validate_freeze_manifest(
                final_manifest_path,
                {final.get("status")},
                FINAL_REVISION_002_REQUIRED_PATHS,
            )
            errors.extend(final_errors)
            errors.extend(validate_sentinel_evidence())
            original_path = SCRIPT_DIR / "EXP3_V2_FREEZE_MANIFEST.json"
            original_errors, original = validate_freeze_manifest(
                original_path,
                {"FROZEN_BEFORE_GENERATION"},
                FINAL_REQUIRED_PATHS,
                historical_boundary=True,
            )
            errors.extend(original_errors)
            baseline_hashes = {
                row["path"]: row["sha256"] for row in original["artifacts"]
            }
            allowed = {
                row["path"]: row for row in final.get("allowed_revision_changes", [])
            }
            observed_changes: set[str] = set()
            for relative, original_hash in baseline_hashes.items():
                observed_hash = sha256_file(ROOT / relative)
                if observed_hash != original_hash:
                    observed_changes.add(relative)
                    change = allowed.get(relative)
                    if (
                        not isinstance(change, dict)
                        or change.get("before_sha256") != original_hash
                        or change.get("after_sha256") != observed_hash
                    ):
                        errors.append(
                            f"unreviewed final-boundary revision change: {relative}"
                        )
            if set(allowed) != observed_changes:
                errors.append("final-boundary revision change set mismatch")
            if plan.get("status") != "FROZEN_BEFORE_GENERATION":
                errors.append("revision 002 must preserve the frozen case plan")
    except Exception as exc:
        errors.append(f"final freeze manifest cannot be read: {exc}")
    errors.extend(validate_history())
    return errors


def validate_sentinel_manifest_row(row: dict[str, str]) -> list[str]:
    errors: list[str] = []
    if row.get("physical_case_id") != ACTIVE_SENTINEL_ID:
        errors.append("sentinel manifest identity mismatch")
    if row.get("attempt") != "0" or row.get("seed") != str(ACTIVE_SENTINEL_SEED):
        errors.append("sentinel manifest attempt/seed mismatch")
    expected_filename = f"{ACTIVE_SENTINEL_ID}__attempt-0.xlsx"
    if row.get("filename") != expected_filename:
        errors.append("sentinel manifest filename mismatch")
    try:
        sampling = float(row["sampling"])
    except (KeyError, TypeError, ValueError):
        errors.append("sentinel manifest sampling is not numeric")
    else:
        if not close(sampling, EXPECTED_SIMULATOR["sampling_interval_h"]):
            errors.append("sentinel manifest sampling mismatch")
    return errors


def sentinel_checks(args: argparse.Namespace) -> list[str]:
    errors: list[str] = []
    plan = json.loads((SCRIPT_DIR / "exp3v2_case_plan.json").read_text())
    descriptor = json.loads(args.sentinel_descriptor.read_text())
    if (
        args.sentinel_descriptor.resolve()
        != (SCRIPT_DIR / "exp3v2_sentinel_case.json").resolve()
    ):
        errors.append("sentinel mode accepts only the frozen descriptor path")
    errors.extend(validate_case_plan(plan))
    errors.extend(validate_sentinel_descriptor(descriptor, plan))
    manifest_errors, _ = validate_freeze_manifest(
        args.harness_manifest, {"HARNESS_FROZEN_FOR_SENTINEL"}, REQUIRED_HARNESS_PATHS
    )
    errors.extend(manifest_errors)
    errors.extend(validate_history())
    harness = json.loads(args.harness_manifest.read_text())
    errors.extend(validate_runtime_directory(harness, args.runtime_dir))
    if (
        args.harness_manifest.resolve()
        != (SCRIPT_DIR / "EXP3_V2_HARNESS_FREEZE_MANIFEST_004.json").resolve()
    ):
        errors.append("sentinel mode accepts only the frozen harness manifest")
    for path in (args.attempt_log, args.manifest, args.data_dir):
        if paths_overlap(path.resolve(), REAL_ROOT) or paths_overlap(
            path.resolve(), OLD_REAL_ROOT
        ):
            errors.append("sentinel path overlaps a real output root")
        if paths_overlap(path.resolve(), args.runtime_dir.resolve()):
            errors.append("sentinel output path overlaps materialized runtime")
    if not path_within(args.attempt_log, args.data_dir.parent) or not path_within(
        args.manifest, args.data_dir.parent
    ):
        errors.append("sentinel artifacts must share one throwaway root")
    log = load_attempt_log(args.attempt_log)
    if len(log["attempts"]) != 1:
        errors.append("sentinel log must contain exactly one attempt")
        return errors
    record = log["attempts"][0]
    if set(record) != ATTEMPT_FIELDS:
        errors.append("sentinel attempt technical-provenance field set mismatch")
    if (
        record.get("physical_case_id") != ACTIVE_SENTINEL_ID
        or record.get("seed") != ACTIVE_SENTINEL_SEED
        or record.get("attempt") != 0
    ):
        errors.append("sentinel attempt identity mismatch")
    for field, value in {**EXPECTED_RUNTIME, **EXPECTED_ATTEMPT_SIMULATOR}.items():
        if record.get(field) != value:
            errors.append(f"sentinel attempt {field} mismatch")
    if record.get("case_plan_hash") != sha256_file(args.sentinel_descriptor):
        errors.append("sentinel attempt descriptor hash mismatch")
    if record.get("generation_script_hash") != sha256_file(
        SCRIPT_DIR / "run_exp3v2_engine.m"
    ):
        errors.append("sentinel attempt engine hash mismatch")
    if not path_within(Path(str(record.get("output_path"))), args.data_dir):
        errors.append("sentinel attempt output path is outside throwaway data dir")
    errors.extend(validate_structural_record(record, "sentinel attempt"))
    rows = load_manifest(args.manifest)
    if len(rows) != 1:
        errors.append("sentinel manifest must contain exactly the sentinel row")
    else:
        errors.extend(validate_sentinel_manifest_row(rows[0]))
        errors.extend(validate_materialized(rows, log, args.data_dir))
    return errors


def post_generation_checks(args: argparse.Namespace) -> list[str]:
    errors: list[str] = []
    plan_path = args.case_plan or SCRIPT_DIR / "exp3v2_case_plan.json"
    plan = json.loads(plan_path.read_text())
    errors.extend(validate_case_plan(plan))
    if plan.get("status") != "FROZEN_BEFORE_GENERATION":
        errors.append("post-generation verification requires final-frozen plan")
    final_path = args.final_manifest or SCRIPT_DIR / "EXP3_V2_FREEZE_MANIFEST_002.json"
    manifest_errors, _ = validate_freeze_manifest(
        final_path, {"FROZEN_BEFORE_GENERATION"}
    )
    errors.extend(manifest_errors)
    if errors:
        return errors
    attempt_log = args.attempt_log or REAL_ROOT / "exp3v2_attempt_log.json"
    manifest = args.manifest or REAL_ROOT / "exp3v2_manifest.csv"
    data_dir = args.data_dir or REAL_ROOT / "mode1"
    if not path_within(attempt_log, REAL_ROOT) or not path_within(data_dir, REAL_ROOT):
        errors.append("real artifact path is outside tep_exp3_v2_heldout")
        return errors
    payload = load_attempt_log(attempt_log)
    schema = json.loads((SCRIPT_DIR / "exp3v2_attempt_log.schema.json").read_text())
    errors.extend(validate_attempt_log(payload, plan, schema))
    rows = load_manifest(manifest)
    expected_ids = [row["physical_case_id"] for row in canonical_cases()]
    if [row["physical_case_id"] for row in rows] != expected_ids:
        errors.append("real manifest is not the exact canonical 30-case set")
    errors.extend(validate_materialized(rows, payload, data_dir))
    errors.extend(validate_history())
    if args.runtime_dir is None:
        errors.append("post-generation verification requires a materialized runtime")
    else:
        final_manifest = json.loads(final_path.read_text())
        errors.extend(validate_runtime_directory(final_manifest, args.runtime_dir))
    return errors


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    modes = parser.add_mutually_exclusive_group()
    modes.add_argument("--pre-freeze", action="store_true")
    modes.add_argument("--sentinel", action="store_true")
    parser.add_argument("--case-plan", type=Path)
    parser.add_argument("--harness-manifest", type=Path)
    parser.add_argument("--final-manifest", type=Path)
    parser.add_argument("--sentinel-descriptor", type=Path)
    parser.add_argument("--attempt-log", type=Path)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--data-dir", type=Path)
    parser.add_argument("--runtime-dir", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        if args.pre_freeze:
            forbidden = (
                args.sentinel_descriptor,
                args.attempt_log,
                args.manifest,
                args.data_dir,
            )
            if any(value is not None for value in forbidden):
                raise ValueError("pre-freeze mode rejects workbook/log/sentinel paths")
            errors = prefreeze_checks(
                args.case_plan or SCRIPT_DIR / "exp3v2_case_plan.json",
                args.harness_manifest
                or SCRIPT_DIR / "EXP3_V2_HARNESS_FREEZE_MANIFEST_004.json",
                args.final_manifest or SCRIPT_DIR / "EXP3_V2_FREEZE_MANIFEST_002.json",
                args.runtime_dir,
            )
            label = "pre-freeze infrastructure"
        elif args.sentinel:
            required = (
                args.sentinel_descriptor,
                args.attempt_log,
                args.manifest,
                args.data_dir,
                args.harness_manifest,
                args.runtime_dir,
            )
            if any(value is None for value in required) or any(
                value is not None for value in (args.case_plan, args.final_manifest)
            ):
                raise ValueError("sentinel mode requires only sentinel-specific paths")
            errors = sentinel_checks(args)
            label = "sentinel"
        else:
            if (
                args.sentinel_descriptor is not None
                or args.harness_manifest is not None
            ):
                raise ValueError("post-generation mode rejects sentinel/harness paths")
            errors = post_generation_checks(args)
            label = "held-out"
    except Exception as exc:
        print(f"FAIL: {exc}", file=sys.stderr)
        return 1
    if errors:
        print("FAIL: Experiment 3 V2 verification failed", file=sys.stderr)
        for error in errors:
            print(f"  - {error}", file=sys.stderr)
        return 1
    print(f"PASS: Experiment 3 V2 {label} verification succeeded.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
