#!/usr/bin/env python3
"""Verify Phase B held-out bytes and workbook structure without signal analysis."""

from __future__ import annotations

import argparse
import csv
import hashlib
import math
import sys
from pathlib import Path
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook


EXPECTED_HEADER = (
    ["Time (h)"]
    + [f"XMEAS-{index}" for index in range(1, 42)]
    + [f"XMV-{index}" for index in range(1, 13)]
)
TIME_TOLERANCE = 1e-10


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def expected_bool(row: dict[str, str], field: str) -> bool:
    value = row[field].strip().lower()
    if value not in {"true", "false"}:
        raise ValueError(f"invalid Boolean in manifest: {field}={row[field]!r}")
    return value == "true"


def close(left: float, right: float) -> bool:
    return math.isclose(left, right, rel_tol=0.0, abs_tol=TIME_TOLERANCE)


def inspect_workbook(path: Path) -> dict[str, object]:
    try:
        with ZipFile(path) as archive:
            xlsx_valid = archive.testzip() is None
    except (BadZipFile, OSError):
        xlsx_valid = False

    if not xlsx_valid:
        raise ValueError("invalid or damaged XLSX ZIP container")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) != 1:
            raise ValueError(f"expected one worksheet, found {workbook.sheetnames!r}")
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header = list(next(rows))
        time_values: list[float] = []
        no_nan = True
        no_inf = True
        finite_numeric = True

        for row_index, row in enumerate(rows, start=2):
            if len(row) != len(EXPECTED_HEADER):
                raise ValueError(
                    f"row {row_index}: expected {len(EXPECTED_HEADER)} cells, "
                    f"found {len(row)}"
                )
            for column_index, value in enumerate(row, start=1):
                if isinstance(value, bool) or not isinstance(value, (int, float)):
                    finite_numeric = False
                    continue
                number = float(value)
                no_nan = no_nan and not math.isnan(number)
                no_inf = no_inf and not math.isinf(number)
                finite_numeric = finite_numeric and math.isfinite(number)
                if column_index == 1:
                    time_values.append(number)

        if len(time_values) < 2:
            raise ValueError("Time column has fewer than two numeric samples")
        intervals = [
            later - earlier
            for earlier, later in zip(time_values, time_values[1:])
        ]
        monotonic = all(interval > 0.0 for interval in intervals)
        sampling_interval_h = intervals[0]
        sampling_constant = all(
            close(interval, sampling_interval_h) for interval in intervals
        )
        complete = (
            sheet.max_row - 1 == 3001
            and close(time_values[0], 0.0)
            and close(time_values[-1], 50.0)
        )

        return {
            "xlsx_valid": xlsx_valid,
            "sheet_name": sheet.title,
            "data_rows": sheet.max_row - 1,
            "columns": sheet.max_column,
            "header_ok": header == EXPECTED_HEADER,
            "finite_numeric": finite_numeric,
            "no_nan": no_nan,
            "no_inf": no_inf,
            "time_start_h": time_values[0],
            "time_end_h": time_values[-1],
            "time_monotonic": monotonic,
            "sampling_interval_h": sampling_interval_h,
            "sampling_interval_min": sampling_interval_h * 60.0,
            "sampling_constant": sampling_constant,
            "trip_or_length_status": (
                "complete_no_early_stop" if complete else "early_stop_or_incomplete"
            ),
        }
    finally:
        workbook.close()


def verify_row(row: dict[str, str], data_dir: Path) -> list[str]:
    errors: list[str] = []
    filename = row["filename"]
    path = data_dir / filename
    if not path.is_file():
        return [f"missing file: {path}"]

    if path.stat().st_size != int(row["size_bytes"]):
        errors.append(
            f"size mismatch: {path.stat().st_size} != {row['size_bytes']}"
        )
    digest = sha256_file(path)
    if digest != row["sha256"]:
        errors.append(f"SHA-256 mismatch: {digest} != {row['sha256']}")

    try:
        observed = inspect_workbook(path)
    except Exception as exc:  # report a per-file integrity failure and continue
        errors.append(str(exc))
        return errors

    integer_fields = ("data_rows", "columns")
    float_fields = (
        "time_start_h",
        "time_end_h",
        "sampling_interval_h",
        "sampling_interval_min",
    )
    boolean_fields = (
        "xlsx_valid",
        "header_ok",
        "finite_numeric",
        "no_nan",
        "no_inf",
        "time_monotonic",
        "sampling_constant",
    )
    text_fields = ("sheet_name", "trip_or_length_status")

    for field in integer_fields:
        if observed[field] != int(row[field]):
            errors.append(f"{field}: {observed[field]} != {row[field]}")
    for field in float_fields:
        if not close(float(observed[field]), float(row[field])):
            errors.append(f"{field}: {observed[field]} != {row[field]}")
    for field in boolean_fields:
        if observed[field] != expected_bool(row, field):
            errors.append(f"{field}: {observed[field]} != {row[field]}")
    for field in text_fields:
        if observed[field] != row[field]:
            errors.append(f"{field}: {observed[field]} != {row[field]}")
    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    script_dir = Path(__file__).resolve().parent
    parser.add_argument(
        "--manifest",
        type=Path,
        default=script_dir / "phase_b_heldout_manifest.csv",
    )
    parser.add_argument(
        "--data-dir",
        type=Path,
        default=script_dir.parents[1] / "tep_heldout" / "mode1",
    )
    parser.add_argument("--quiet", action="store_true")
    args = parser.parse_args()

    with args.manifest.open(newline="", encoding="utf-8") as stream:
        rows = list(csv.DictReader(stream))
    filenames = [row["filename"] for row in rows]
    if len(rows) != 15 or len(set(filenames)) != 15:
        print(
            "ERROR: manifest must contain exactly 15 unique workbooks",
            file=sys.stderr,
        )
        return 1

    failures = 0
    for row in rows:
        errors = verify_row(row, args.data_dir)
        if errors:
            failures += 1
            print(f"FAIL {row['case_id']} {row['filename']}", file=sys.stderr)
            for error in errors:
                print(f"  - {error}", file=sys.stderr)
        elif not args.quiet:
            print(f"OK   {row['case_id']} {row['filename']}")

    if failures:
        print(f"Integrity verification failed for {failures}/15 files.", file=sys.stderr)
        return 1
    print("Integrity verification succeeded for all 15 frozen workbooks.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
